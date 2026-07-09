# /// script
# dependencies = [
#   "flask",
#   "flask-cors",
#   "openpyxl",
#   "requests",
# ]
# ///

"""
MuHub Excel Server & LINE Webhook
==================================
รัน: python muhub_excel_server.py
ทำงานเป็น local API ที่:
1. รับข้อมูลโดยตรงจาก booking app (index.html) บันทึกลงชีต Bookings
2. รับข้อมูลข้อความลูกค้าจาก LINE OA @muhub ผ่าน Webhook บันทึกลงชีต Cust
"""

from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import sys
import json
import base64
import hmac
import hashlib
import requests
import time
import threading
import logging
from logging.handlers import RotatingFileHandler
from collections import defaultdict, deque

app = Flask(__name__)
CORS(app)

# ── Logging: บันทึกทุกอย่างที่เคย print() ลงไฟล์ log แบบ rotate อัตโนมัติ ──────
# ไฟล์นี้ใช้ print() เป็น logger อยู่แล้วเกือบ 100 จุด แทนที่จะไปแก้ print() ทีละจุด
# (เสี่ยงพลาด/ตกหล่น) เราแค่ "ดักฟัง" stdout/stderr แล้วเขียนสำเนาออกไฟล์ log ด้วย
# ทำให้ log เดิมทั้งหมดถูกเก็บถาวรลงไฟล์ (ไม่หายไปเมื่อปิด terminal/รันเป็น background)
BASE_DIR_FOR_LOG = os.path.dirname(os.path.abspath(__file__))
LOG_DIR = os.path.join(BASE_DIR_FOR_LOG, "logs")
os.makedirs(LOG_DIR, exist_ok=True)

_excel_server_logger = logging.getLogger("muhub_excel_server")
_excel_server_logger.setLevel(logging.INFO)
_excel_server_logger.propagate = False
if not _excel_server_logger.handlers:
    _log_handler = RotatingFileHandler(
        os.path.join(LOG_DIR, "excel_server.log"), maxBytes=5 * 1024 * 1024, backupCount=5, encoding="utf-8"
    )
    _log_handler.setFormatter(logging.Formatter("%(asctime)s %(message)s"))
    _excel_server_logger.addHandler(_log_handler)

class _StreamToLogger:
    """ส่งต่อทุกบรรทัดที่เขียนไปยัง stdout/stderr (เช่นจาก print()) เข้า logger ไฟล์ด้วย"""
    def __init__(self, logger, level, original_stream):
        self.logger = logger
        self.level = level
        self.original_stream = original_stream
        self._buffer = ""

    def write(self, message):
        self.original_stream.write(message)
        self._buffer += message
        while "\n" in self._buffer:
            line, self._buffer = self._buffer.split("\n", 1)
            if line.strip():
                self.logger.log(self.level, line)

    def flush(self):
        self.original_stream.flush()

    def isatty(self):
        return hasattr(self.original_stream, 'isatty') and self.original_stream.isatty()

    def fileno(self):
        if hasattr(self.original_stream, 'fileno'):
            return self.original_stream.fileno()
        raise OSError("Stream has no fileno")

sys.stdout = _StreamToLogger(_excel_server_logger, logging.INFO, sys.stdout)
sys.stderr = _StreamToLogger(_excel_server_logger, logging.ERROR, sys.stderr)

# ── Basic error monitoring: จับ exception ที่ไม่ได้ handle ไว้ทุกจุด ─────────
# ถ้าไม่มีตัวนี้ exception ที่หลุดจาก route ใดๆ จะทำให้ client เห็น stack trace เต็มๆ
# (เสี่ยงข้อมูลรั่ว) ตัวนี้ log รายละเอียด error เต็มๆ ลงไฟล์ แล้วตอบ client แบบสุภาพ
@app.errorhandler(Exception)
def handle_unhandled_exception(e):
    _excel_server_logger.exception(f"[UnhandledError] {request.method} {request.path}: {e}")
    return jsonify({"success": False, "error": "internal_error", "message": "เกิดข้อผิดพลาดที่ไม่คาดคิดในระบบ กรุณาลองใหม่อีกครั้ง"}), 500

# ── โหลดการตั้งค่าจาก config.json ─────────────────────────────────────────
CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
EXCEL_PATH = r"G:\My Drive\Customer\MuHub_Customer_DB.xlsx"
LINE_CHANNEL_ACCESS_TOKEN = ""
LINE_CHANNEL_SECRET = ""
LINE_ADMIN_USER_ID = ""
WEB_APP_URL = "http://localhost:3737/"
GOOGLE_SCRIPT_URL = "https://script.google.com/macros/s/AKfycbwUkfB7te_1-o_u_2ZroBzpuACWmFYd8XS5JN3iv9rItvql-887wQw95SIIQBXWUQumPA/exec"
PORT = 5001
SLIPOK_API_KEY = ""
SLIPOK_BRANCH_ID = ""
APP_API_TOKEN = ""

if os.path.exists(CONFIG_PATH):
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            config = json.load(f)
            EXCEL_PATH = config.get("EXCEL_PATH", EXCEL_PATH)
            LINE_CHANNEL_ACCESS_TOKEN = config.get("LINE_CHANNEL_ACCESS_TOKEN", "")
            LINE_CHANNEL_SECRET = config.get("LINE_CHANNEL_SECRET", "")
            LINE_ADMIN_USER_ID = config.get("LINE_ADMIN_USER_ID", "")
            WEB_APP_URL = config.get("WEB_APP_URL", "http://localhost:3737/")
            GOOGLE_SCRIPT_URL = config.get("GOOGLE_SCRIPT_URL", GOOGLE_SCRIPT_URL)
            PORT = config.get("PORT", PORT)
            SLIPOK_API_KEY = config.get("SLIPOK_API_KEY", "")
            SLIPOK_BRANCH_ID = config.get("SLIPOK_BRANCH_ID", "")
            APP_API_TOKEN = config.get("APP_API_TOKEN", "")
            print(f"[Config] Loaded settings from config.json")
            print(f"[Config] Excel Path: {EXCEL_PATH}")
            print(f"[Config] LINE Token configured: {bool(LINE_CHANNEL_ACCESS_TOKEN)}")
            print(f"[Config] LINE Admin User ID configured: {bool(LINE_ADMIN_USER_ID)}")
            print(f"[Config] SlipOK configured: {bool(SLIPOK_API_KEY)}")
            print(f"[Config] App API Token configured: {bool(APP_API_TOKEN)}")
            print(f"[Config] Web App URL: {WEB_APP_URL}")
            print(f"[Config] Google Script URL: {GOOGLE_SCRIPT_URL}")
    except Exception as e:
        print(f"[Config] Warning: Failed to load config.json: {e}")
else:
    # สร้างไฟล์ตั้งค่าเริ่มต้นหากยังไม่มี
    try:
        default_config = {
            "LINE_CHANNEL_ACCESS_TOKEN": "",
            "LINE_CHANNEL_SECRET": "",
            "LINE_ADMIN_USER_ID": "",
            "WEB_APP_URL": "http://localhost:3737/",
            "EXCEL_PATH": EXCEL_PATH,
            "PORT": PORT
        }
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(default_config, f, ensure_ascii=False, indent=2)
        print(f"[Config] Created template config.json")
    except Exception as e:
        print(f"[Config] Warning: Failed to create config.json: {e}")

# ── การตั้งค่าชีต Excel ───────────────────────────────────────────────────
HEADERS = [
    "Timestamp", "Queue ID", "Customer Name", "LINE ID",
    "Astrologer", "Service", "Date", "Time Slot", "Questions",
    "Birth Date", "Birth Hour", "Birth Min",
    "Birth Country", "Birth City", "GCal Event ID", "Status"
]

CUST_HEADERS = [
    "Customer ID", "วันที่บันทึก", "Display Name", "LINE User ID", "สถานะ", 
    "วันที่จอง", "ชื่อลูกค้า", "Line ID", "วันเกิด (พ.ศ.)", "เวลาเกิด", 
    "ประเทศ", "จังหวัด", "แพคเกจ/บริการ", "หมอดู", "วันนัดหมาย", 
    "ช่วงเวลานัด", "คำถามพิเศษ", "เลขอ้างอิง", "สถานะการชำระ", "ประเภทลูกค้า", 
    "จำนวนครั้งที่จอง", "หมายเหตุ"
]

QUEUE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pending_cust_writes.json")
PROFILE_CACHE = {}

# ── โปรโมโค้ด: เก็บรายชื่อ Line ID ที่ใช้สิทธิ์ไปแล้ว (server-side, กันการใช้ซ้ำผ่าน localStorage) ──
PROMO_USED_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "promo_used.json")
promo_lock = threading.Lock()  # กัน race condition เวลามีคำขอ redeem เข้ามาพร้อมกัน

# ── การป้องกัน Endpoint ที่เขียนข้อมูล (Token + Rate Limit) ───────────────────
# เนื่องจาก server นี้ถูก tunnel ออกอินเทอร์เน็ต (สำหรับ LINE webhook) endpoint
# ที่ "เขียน" ข้อมูลจริง (/api/save-customer, /api/redeem-promo) จึงมีคนภายนอก
# ยิง request ปลอมเข้ามาได้โดยตรง ถ้าไม่มีการป้องกันใดๆ เลย
#
# หมายเหตุสำคัญ: เว็บแอปนี้เป็น static frontend ล้วน (ไม่มี login/session) ดังนั้น
# APP_API_TOKEN ที่ฝังไว้ใน booking.js จะ "มองเห็นได้" ผ่าน view-source เสมอ
# มันจึงช่วยกัน bot/สคริปต์กราดสุ่มยิงได้ ไม่ใช่การยืนยันตัวตนที่ปลอดภัย 100%
# ต่อผู้โจมตีที่ตั้งใจอ่านโค้ด JS ก่อน — ควบคู่กับ rate limit ด้านล่างจะช่วยลด
# ความเสียหายจากการยิงถล่ม (spam) ได้ในระดับที่เหมาะกับสเกลปัจจุบัน
_rate_limit_buckets = defaultdict(deque)
_rate_limit_lock = threading.Lock()
RATE_LIMIT_MAX_REQUESTS = 8      # จำนวนครั้งสูงสุดต่อหน้าต่างเวลา
RATE_LIMIT_WINDOW_SECONDS = 60   # หน้าต่างเวลา (วินาที)

def _get_client_ip():
    """ดึง IP จริงของ client โดยรองรับกรณีอยู่หลัง tunnel/reverse proxy"""
    forwarded = request.headers.get("X-Forwarded-For", "")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.remote_addr or "unknown"

def check_rate_limit(bucket_key):
    """คืนค่า True ถ้ายังไม่เกิน rate limit, False ถ้าเกินแล้ว (ควรปฏิเสธ request)"""
    now = time.time()
    with _rate_limit_lock:
        q = _rate_limit_buckets[bucket_key]
        while q and now - q[0] > RATE_LIMIT_WINDOW_SECONDS:
            q.popleft()
        if len(q) >= RATE_LIMIT_MAX_REQUESTS:
            return False
        q.append(now)
        return True

def require_app_token_and_rate_limit(endpoint_name):
    """
    เช็ค token (ถ้าตั้งค่า APP_API_TOKEN ไว้) และ rate limit ต่อ IP
    เรียกใช้ที่ต้นทุกฟังก์ชันของ endpoint ที่ "เขียน" ข้อมูล
    คืนค่า None ถ้าผ่าน หรือ (response, status_code) ถ้าควรปฏิเสธ
    """
    client_ip = _get_client_ip()

    if not check_rate_limit(f"{endpoint_name}:{client_ip}"):
        print(f"[Security] Rate limit exceeded for {endpoint_name} from {client_ip}")
        return jsonify({"success": False, "error": "rate_limited", "message": "ส่งคำขอถี่เกินไป กรุณาลองใหม่ภายหลัง"}), 429

    if APP_API_TOKEN:
        provided = request.headers.get("X-App-Token", "")
        if provided != APP_API_TOKEN:
            print(f"[Security] Invalid/missing app token for {endpoint_name} from {client_ip}")
            return jsonify({"success": False, "error": "unauthorized"}), 401

    return None

# ── ป้องกัน Formula/CSV Injection ──────────────────────────────────────────
# ข้อมูลทุกฟิลด์ที่มาจากลูกค้า (เว็บฟอร์ม หรือแชท LINE) ถือเป็น "ข้อมูลที่ไม่น่าเชื่อถือ"
# หากลูกค้า (หรือบอท) ตั้งใจพิมพ์ค่าที่ขึ้นต้นด้วย =, +, -, @ เช่น
# "=IMPORTXML(...)" หรือ "=cmd|'/c calc'!A1" แล้วแอดมินเปิดไฟล์ Excel ขึ้นมา
# Excel/openpyxl อาจตีความค่านั้นเป็นสูตรและรันโค้ดอันตรายได้ (ช่องโหว่ที่รู้จักกันในชื่อ
# CSV/Formula Injection) ฟังก์ชันนี้ป้องกันโดยการเติมเครื่องหมาย ' นำหน้า
# ค่าที่ขึ้นต้นด้วยอักขระอันตราย เพื่อบังคับให้ openpyxl เก็บเป็นข้อความล้วน ไม่ใช่สูตร
_DANGEROUS_LEADING_CHARS = ("=", "+", "-", "@", "\t", "\r")

def sanitize_excel_value(val):
    """แปลงค่าที่จะเขียนลง Excel cell ให้ปลอดภัยจาก formula injection"""
    if isinstance(val, str) and val.startswith(_DANGEROUS_LEADING_CHARS):
        return "'" + val
    return val

# ── ฟังก์ชันช่วยสำหรับไฟล์ Excel ───────────────────────────────────────────
def get_or_create_workbook():
    """โหลดหรือสร้าง workbook สำหรับบันทึกข้อมูลจองคิวหลัก"""
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        # ตรวจสอบหาชีต Bookings หรือใช้ active sheet
        if "Bookings" in wb.sheetnames:
            ws = wb["Bookings"]
        else:
            ws = wb.active
            if ws.max_row == 0 or ws.cell(1, 1).value is None:
                ws.title = "Bookings"
                ws.append(HEADERS)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bookings"
        ws.append(HEADERS)
        for col in range(1, len(HEADERS) + 1):
            ws.cell(1, col).font = openpyxl.styles.Font(bold=True)
    return wb

# ── หาคอลัมน์ของชีต Bookings จากชื่อ header แทนเลข hardcode ────────────────
# HEADERS ที่ประกาศไว้ด้านบนมี 16 คอลัมน์ (GCal Event ID = คอลัมน์ 15, Status = 16)
# แต่โค้ดส่วน auto-confirm ที่เขียนไว้ก่อนหน้านี้กลับอ่าน/เขียนที่คอลัมน์ 20/21
# ซึ่งแปลว่าถ้าไฟล์ Excel จริงของแอดมิน (G:\My Drive\...) ไม่เคยมีคอลัมน์เพิ่มเติม
# ถูกแทรกไว้ ระบบ auto-confirm ของคิวที่จองตรงจากหน้าเว็บจะอ่าน/เขียนคอลัมน์ผิด
# ตำแหน่งไปเงียบๆ โดยไม่มีใครรู้ ฟังก์ชันนี้แก้ปัญหาด้วยการหาคอลัมน์จาก "ชื่อ header"
# ในแถวที่ 1 ก่อนเสมอ (ยืดหยุ่นตามไฟล์จริงไม่ว่าแอดมินจะสลับ/แทรกคอลัมน์อย่างไร)
# ถ้าหาไม่เจอจริงๆ ค่อย fallback ไปที่เลข 20/21 เดิม พร้อม print แจ้งเตือนให้สังเกตเห็น
def resolve_bookings_columns(ws):
    """คืนค่า (col_gcal_event_id, col_status) ของชีต Bookings โดยหาจากชื่อ header ก่อน"""
    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val:
            header_map[str(val).strip()] = col_idx

    col_gcal = header_map.get("GCal Event ID")
    col_status = header_map.get("Status")

    if not col_gcal or not col_status:
        print(
            "[Bookings] คำเตือน: ไม่พบคอลัมน์ 'GCal Event ID' หรือ 'Status' จากชื่อ header "
            "แถวที่ 1 ของชีต Bookings — ใช้เลขคอลัมน์ fallback (20/21) ไปก่อน "
            "กรุณาตรวจสอบว่า header แถวแรกของไฟล์ Excel จริงตรงกับที่ระบบคาดไว้"
        )

    return col_gcal or 20, col_status or 21

def get_cust_sheet(wb):
    """ดึงหรือสร้างชีต Cust สำหรับฐานข้อมูลลูกค้า LINE"""
    if "Cust" in wb.sheetnames:
        ws = wb["Cust"]
    else:
        ws = wb.create_sheet("Cust")
        ws.append(CUST_HEADERS)
        for col in range(1, len(CUST_HEADERS) + 1):
            ws.cell(1, col).font = openpyxl.styles.Font(bold=True)
    return ws

def save_workbook_safe(wb, path, retries=5, delay=1.5):
    """พยายามบันทึกไฟล์ Excel ป้องกันการล้มเหลวหากไฟล์โดนล็อกอยู่"""
    for i in range(retries):
        try:
            wb.save(path)
            print(f"[Excel] Saved workbook successfully to {path}")
            return True
        except PermissionError:
            print(f"[Excel] Permission denied (file locked). Retry {i+1}/{retries} in {delay}s...")
            time.sleep(delay)
        except Exception as e:
            print(f"[Excel] Error saving workbook: {e}")
            raise e
    print(f"[Excel] Error: Failed to save workbook after {retries} retries.")
    return False

# ── ฟังก์ชันบริการคิวสำรองกรณีไฟล์โดนล็อก ─────────────────────────────────────
def queue_pending_write(item):
    """บันทึกรายการเขียนที่ล้มเหลวลงไฟล์คิวสำรองป้องกันข้อมูลสูญหาย"""
    queue = []
    if os.path.exists(QUEUE_FILE):
        try:
            with open(QUEUE_FILE, "r", encoding="utf-8") as f:
                queue = json.load(f)
        except Exception:
            queue = []
    queue.append(item)
    try:
        with open(QUEUE_FILE, "w", encoding="utf-8") as f:
            json.dump(queue, f, ensure_ascii=False, indent=2)
        print(f"[Queue] Queued write event. Total pending: {len(queue)}")
    except Exception as e:
        print(f"[Queue] Error writing queue file: {e}")

def process_pending_queue(wb, ws):
    """เขียนข้อมูลจากคิวสำรองลงในชีต Cust"""
    if not os.path.exists(QUEUE_FILE):
        return False
    
    queue = []
    try:
        with open(QUEUE_FILE, "r", encoding="utf-8") as f:
            queue = json.load(f)
    except Exception:
        return False
        
    if not queue:
        return False
        
    print(f"[Queue] Processing {len(queue)} pending queue items...")
    successful_indices = []
    
    for idx, item in enumerate(queue):
        try:
            data = item.get("data")
            append_only_user_id = item.get("append_only_user_id")
            
            if append_only_user_id:
                row_idx = append_message_to_existing_user(ws, append_only_user_id, data.get("notes", ""))
                if not row_idx:
                    write_data_to_sheet(ws, data)
            else:
                write_data_to_sheet(ws, data)
                
            successful_indices.append(idx)
        except Exception as e:
            print(f"[Queue] Failed to process queue item {idx}: {e}")
            break # หยุดเขียนเพื่อรักษาระเบียบข้อมูล
            
    if successful_indices:
        new_queue = [item for i, item in enumerate(queue) if i not in successful_indices]
        try:
            if new_queue:
                with open(QUEUE_FILE, "w", encoding="utf-8") as f:
                    json.dump(new_queue, f, ensure_ascii=False, indent=2)
            else:
                if os.path.exists(QUEUE_FILE):
                    os.remove(QUEUE_FILE)
            print(f"[Queue] Successfully cleared {len(successful_indices)} items. Remaining: {len(new_queue)}")
            return True
        except Exception as e:
            print(f"[Queue] Error updating queue file: {e}")
    return False

# ── ฟังก์ชันการจัดระเบียบและเขียนข้อมูลในชีต Cust ──────────────────────────────
def write_data_to_sheet(ws, data):
    """ค้นหาแถวว่างและเขียนชุดข้อมูลใหม่ลงในชีต Cust"""
    row_idx = 3
    while True:
        c_val = ws.cell(row=row_idx, column=3).value  # Display Name
        d_val = ws.cell(row=row_idx, column=4).value  # LINE User ID
        if c_val is None and d_val is None:
            break
        row_idx += 1
        
    print(f"[Excel] Writing new row to line {row_idx}")
    
    ws.cell(row=row_idx, column=1).value = f'=IF(C{row_idx}<>"",TEXT(ROW()-3,"0000"),"")'
    ws.cell(row=row_idx, column=2).value = sanitize_excel_value(data.get("timestamp", datetime.now().strftime("%d/%m/%Y %H:%M")))
    ws.cell(row=row_idx, column=3).value = sanitize_excel_value(data.get("display_name", ""))
    ws.cell(row=row_idx, column=4).value = sanitize_excel_value(data.get("user_id", ""))
    ws.cell(row=row_idx, column=5).value = sanitize_excel_value(data.get("status", "ใหม่"))
    ws.cell(row=row_idx, column=6).value = sanitize_excel_value(data.get("booking_date", ""))
    ws.cell(row=row_idx, column=7).value = sanitize_excel_value(data.get("customer_name", ""))
    ws.cell(row=row_idx, column=8).value = sanitize_excel_value(data.get("line_id", ""))
    ws.cell(row=row_idx, column=9).value = sanitize_excel_value(data.get("birth_date", ""))
    ws.cell(row=row_idx, column=10).value = sanitize_excel_value(data.get("birth_time", ""))
    ws.cell(row=row_idx, column=11).value = sanitize_excel_value(data.get("birth_country", ""))
    ws.cell(row=row_idx, column=12).value = sanitize_excel_value(data.get("birth_city", ""))
    ws.cell(row=row_idx, column=13).value = sanitize_excel_value(data.get("package", ""))
    ws.cell(row=row_idx, column=14).value = sanitize_excel_value(data.get("astrologer", ""))
    ws.cell(row=row_idx, column=15).value = sanitize_excel_value(data.get("booking_date_appointment", ""))
    ws.cell(row=row_idx, column=16).value = sanitize_excel_value(data.get("booking_slot", ""))
    ws.cell(row=row_idx, column=17).value = sanitize_excel_value(data.get("questions", ""))
    ws.cell(row=row_idx, column=18).value = sanitize_excel_value(data.get("queue_id", ""))
    ws.cell(row=row_idx, column=19).value = sanitize_excel_value(data.get("payment_status", ""))
    ws.cell(row=row_idx, column=20).value = f'=IF(U{row_idx}>1,"ลูกค้าเก่า","ลูกค้าใหม่")'
    ws.cell(row=row_idx, column=21).value = f'=COUNTIF(D:D, D{row_idx})'
    ws.cell(row=row_idx, column=22).value = sanitize_excel_value(data.get("notes", ""))
    
    # Auto-fit columns
    for col_num in range(1, len(CUST_HEADERS) + 1):
        col_letter = get_column_letter(col_num)
        max_len = max((len(str(cell.value or "")) for cell in ws[col_letter]), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

def append_message_to_existing_user(ws, user_id, message_text):
    """นำข้อความไปต่อท้ายในหมายเหตุ (คอลัมน์ V) ของแถวล่าสุดของผู้ใช้นั้น"""
    matched_rows = []
    for r in range(3, ws.max_row + 1):
        d_val = ws.cell(row=r, column=4).value
        if d_val == user_id:
            matched_rows.append(r)
            
    if matched_rows:
        latest_row = matched_rows[-1]
        cell = ws.cell(row=latest_row, column=22) # คอลัมน์ V (หมายเหตุ)
        old_val = cell.value
        timestamp = datetime.now().strftime("%d/%m %H:%M")
        note_line = f"[{timestamp}] ลูกค้า: {message_text}"
        if old_val:
            cell.value = sanitize_excel_value(f"{old_val}\n{note_line}")
        else:
            cell.value = sanitize_excel_value(note_line)
        print(f"[Excel] Appended note to existing user at row {latest_row}")
        return latest_row
    return None

def get_pending_cust_sheet(wb):
    """ดึงหรือสร้างชีต PendingCust สำหรับรายการรอลูกดวงชำระเงิน"""
    if "PendingCust" in wb.sheetnames:
        ws = wb["PendingCust"]
    else:
        ws = wb.create_sheet("PendingCust")
        ws.append(CUST_HEADERS)
        for col in range(1, len(CUST_HEADERS) + 1):
            ws.cell(1, col).font = openpyxl.styles.Font(bold=True)
    return ws

def process_data_write(data, append_only_user_id=None, target_sheet="Cust"):
    """เรียกประมวลผลการจัดเขียนไฟล์ Excel ในลักษณะควบคุมความปลอดภัยและจัดการความทนทานต่อการล็อค"""
    try:
        wb = load_workbook(EXCEL_PATH) if os.path.exists(EXCEL_PATH) else openpyxl.Workbook()
        
        if target_sheet == "PendingCust":
            ws = get_pending_cust_sheet(wb)
        else:
            ws = get_cust_sheet(wb)
            
        # เขียนคิวค้างส่งก่อน (เฉพาะของชีตหลัก)
        if target_sheet == "Cust":
            process_pending_queue(wb, ws)
        
        if append_only_user_id:
            row_idx = append_message_to_existing_user(ws, append_only_user_id, data.get("notes", ""))
            if not row_idx:
                # สำรองในกรณีไม่พบแถวเก่า ให้เขียนใหม่ทั้งหมด
                write_data_to_sheet(ws, data)
        else:
            write_data_to_sheet(ws, data)
            
        success = save_workbook_safe(wb, EXCEL_PATH)
        if not success:
            # หากล็อคอยู่ ให้ลงคิวสำรอง (เฉพาะของชีตหลัก)
            if target_sheet == "Cust":
                queue_data = {
                    "data": data,
                    "append_only_user_id": append_only_user_id
                }
                queue_pending_write(queue_data)
            return False
        return True
    except Exception as e:
        print(f"[Excel] Fatal error processing write: {e}")
        if target_sheet == "Cust":
            queue_data = {
                "data": data,
                "append_only_user_id": append_only_user_id
            }
            queue_pending_write(queue_data)
        return False

def parse_date_slot_to_iso(date_str, slot_str):
    """
    แปลงวันที่และช่วงเวลาให้เป็น ISO format (เวลาประเทศไทย UTC+7) สำหรับส่ง Google Calendar
    เช่น:
      date_str = "2026-06-30" หรือ "30/06/2026"
      slot_str = "10:00 - 11:00" หรือ "10:00-11:00"
    ส่งกลับ: (start_iso_str, end_iso_str)
    """
    import re
    date_str = date_str.strip()
    
    # แบบ YYYY-MM-DD
    match_yyyy = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$", date_str)
    # แบบ DD/MM/YYYY
    match_dd = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$", date_str)
    
    yyyy, mm, dd = None, None, None
    
    if match_yyyy:
        yyyy, mm, dd = match_yyyy.groups()
    elif match_dd:
        dd, mm, yyyy = match_dd.groups()
    else:
        # ลองสแกนหาตัวเลขทั่วไป
        nums = re.findall(r"\d+", date_str)
        if len(nums) >= 3:
            if len(nums[0]) == 4:
                yyyy, mm, dd = nums[0], nums[1], nums[2]
            elif len(nums[2]) == 4:
                dd, mm, yyyy = nums[0], nums[1], nums[2]
                
    if not yyyy or not mm or not dd:
        raise ValueError("รูปแบบวันที่ไม่ถูกต้อง (ตัวอย่าง: 2026-06-30 หรือ 30/06/2026)")
        
    # ปรับปี พ.ศ. เป็น ค.ศ. (ถ้าปีเกิน 2500)
    year_val = int(yyyy)
    if year_val > 2500:
        year_val -= 543
        
    date_normalized = f"{year_val:04d}-{int(mm):02d}-{int(dd):02d}"
    
    # 2. ปรับรูปแบบเวลา
    slot_str = slot_str.replace(" ", "").replace(".", ":")
    times = re.findall(r"\d{2}:\d{2}", slot_str)
    
    if len(times) == 2:
        start_time, end_time = times[0], times[1]
    elif len(times) == 1:
        start_time = times[0]
        h, m = map(int, start_time.split(":"))
        end_time = f"{(h+1)%24:02d}:{m:02d}"
    else:
        nums = re.findall(r"\d+", slot_str)
        if len(nums) >= 2:
            start_time = f"{int(nums[0]):02d}:00"
            end_time = f"{int(nums[1]):02d}:00"
        else:
            raise ValueError("รูปแบบเวลานัดไม่ถูกต้อง (ตัวอย่าง: 10:00 - 11:00)")
            
    # ประกอบเป็น ISO String ไทม์โซนประเทศไทย UTC+7
    start_iso = f"{date_normalized}T{start_time}:00+07:00"
    end_iso = f"{date_normalized}T{end_time}:00+07:00"
    
    return start_iso, end_iso

def format_date_to_dd_mm_yyyy(date_str):
    """
    แปลงวันที่ที่ลูกค้ากรอกมาให้อยู่ในรูปแบบ DD/MM/YYYY (ปี ค.ศ.) เสมอ
    """
    import re
    date_str = date_str.strip()
    match_yyyy = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$", date_str)
    match_dd = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$", date_str)
    
    yyyy, mm, dd = None, None, None
    if match_yyyy:
        yyyy, mm, dd = match_yyyy.groups()
    elif match_dd:
        dd, mm, yyyy = match_dd.groups()
    else:
        nums = re.findall(r"\d+", date_str)
        if len(nums) >= 3:
            if len(nums[0]) == 4:
                yyyy, mm, dd = nums[0], nums[1], nums[2]
            elif len(nums[2]) == 4:
                dd, mm, yyyy = nums[0], nums[1], nums[2]
                
    if not yyyy or not mm or not dd:
        return date_str  # คืนค่าเดิมถ้าแปลงไม่ได้
        
    year_val = int(yyyy)
    if year_val > 2500:
        year_val -= 543
        
    return f"{int(dd):02d}/{int(mm):02d}/{year_val:04d}"

def move_row_to_confirmed_sheet(wb, pending_ws, r):
    """
    ย้ายแถวข้อมูลการจองจากชีต PendingCust ไปยัง Cust เมื่อยืนยันการชำระเงินเรียบร้อยแล้ว
    """
    if "Cust" not in wb.sheetnames:
        ws_cust = wb.create_sheet("Cust")
        # เขียน headers
        for col_num, header in enumerate(CUST_HEADERS, 1):
            ws_cust.cell(row=1, column=col_num).value = header
    else:
        ws_cust = wb["Cust"]
        
    # หาแถวใหม่ใน Cust
    cust_row_idx = 3
    while True:
        c_val = ws_cust.cell(row=cust_row_idx, column=3).value
        d_val = ws_cust.cell(row=cust_row_idx, column=4).value
        if c_val is None and d_val is None:
            break
        cust_row_idx += 1
        
    # คัดลอกข้อมูลทุกคอลัมน์จาก PendingCust (แถว r) ไปยัง Cust (แถว cust_row_idx)
    for col_idx in range(1, 23):
        val = pending_ws.cell(row=r, column=col_idx).value
        ws_cust.cell(row=cust_row_idx, column=col_idx).value = val
        
    # อัปเดตสูตรอัตโนมัติใน Cust
    ws_cust.cell(row=cust_row_idx, column=1).value = f'=IF(C{cust_row_idx}<>"",TEXT(ROW()-3,"0000"),"")'
    ws_cust.cell(row=cust_row_idx, column=20).value = f'=IF(U{cust_row_idx}>1,"ลูกค้าเก่า","ลูกค้าใหม่")'
    ws_cust.cell(row=cust_row_idx, column=21).value = f'=COUNTIF(D:D, D{cust_row_idx})'
    
    # ลบแถวออกจาก PendingCust
    pending_ws.delete_rows(r)

def confirm_and_gcal_booking(queue_id):
    """
    ค้นหารหัสคิวใน Excel:
    1. ค้นหาในชีต PendingCust (คิวรอชำระเงินจากแชท)
    2. ค้นหาในชีต Bookings (คิวจากเว็บแอป)
    หากพบ ให้ทำการตรวจสอบคิวใน Google Calendar และลงล็อก
    จากนั้นเขียนลงชีต Cust หลัก (หรืออัปเดต Bookings) เป็นขั้นตอนสุดท้าย และแจ้งเตือนลูกค้า
    """
    if not queue_id:
        return {"success": False, "message": "กรุณาระบุรหัสคิว เช่น ยืนยัน Q-1234"}
        
    queue_id_clean = str(queue_id).strip().upper()
    
    if not os.path.exists(EXCEL_PATH):
        return {"success": False, "message": "ไม่พบไฟล์ฐานข้อมูล Excel"}
        
    try:
        wb = load_workbook(EXCEL_PATH)
        target_sheet = None
        target_row = None
        
        # 1. ค้นหาในชีต PendingCust (รอยืนยัน)
        if "PendingCust" in wb.sheetnames:
            ws = wb["PendingCust"]
            for r in range(3, ws.max_row + 1):
                val = ws.cell(row=r, column=18).value
                if val and str(val).strip().upper() == queue_id_clean:
                    target_sheet = "PendingCust"
                    target_row = r
                    break
                    
        # 2. ค้นหาในชีต Bookings (จากเว็บแอป)
        if not target_sheet and "Bookings" in wb.sheetnames:
            ws = wb["Bookings"]
            for r in range(2, ws.max_row + 1):
                val = ws.cell(row=r, column=2).value
                if val and str(val).strip().upper() == queue_id_clean:
                    target_sheet = "Bookings"
                    target_row = r
                    break
                    
        # หากไม่พบในทั้ง 2 ชีต อาจจะเคยอนุมัติไปแล้วในชีต Cust หลัก
        if not target_sheet:
            if "Cust" in wb.sheetnames:
                ws_cust = wb["Cust"]
                for r in range(3, ws_cust.max_row + 1):
                    val = ws_cust.cell(row=r, column=18).value
                    if val and str(val).strip().upper() == queue_id_clean:
                        return {"success": False, "message": f"คิว {queue_id_clean} ได้ยืนยันและบันทึกลงฐานข้อมูลหลักเรียบร้อยแล้วก่อนหน้านี้ค่ะ"}
            return {"success": False, "message": f"ไม่พบรหัสคิว {queue_id_clean} ในตารางข้อมูลรอยืนยันค่ะ"}
            
        ws = wb[target_sheet]
        r = target_row
        
        # ดึงข้อมูลเพื่อตรวจปฏิทิน
        if target_sheet == "PendingCust":
            customer_name = ws.cell(row=r, column=7).value or "ลูกค้า"
            line_id = ws.cell(row=r, column=8).value or ""
            user_id = ws.cell(row=r, column=4).value or ""
            package = ws.cell(row=r, column=13).value or ""
            astrologer = ws.cell(row=r, column=14).value or ""
            booking_date = ws.cell(row=r, column=15).value or ""
            booking_slot = ws.cell(row=r, column=16).value or ""
            questions = ws.cell(row=r, column=17).value or ""
        else: # Bookings
            customer_name = ws.cell(row=r, column=3).value or "ลูกค้า"
            line_id = ws.cell(row=r, column=4).value or ""
            user_id = find_user_id_by_line_id(line_id) or ""
            package = ws.cell(row=r, column=6).value or ""
            astrologer = ws.cell(row=r, column=5).value or ""
            booking_date = ws.cell(row=r, column=7).value or ""
            booking_slot = ws.cell(row=r, column=8).value or ""
            questions = ws.cell(row=r, column=9).value or ""
            
        # 3. จอง Google Calendar
        cal_result = check_and_block_calendar_event(
            date_str=str(booking_date),
            slot_str=str(booking_slot),
            name=str(customer_name),
            line_id=str(line_id),
            service=str(package),
            astrologer=str(astrologer),
            questions=str(questions)
        )
        
        if not cal_result.get("success"):
            return {
                "success": False, 
                "message": f"❌ ตารางเวลาชนกัน หรือเกิดข้อผิดพลาดในการล็อกปฏิทิน: {cal_result.get('message')}"
            }
            
        gcal_event_id = cal_result.get("eventId")
        
        # 4. ดำเนินการย้าย/อัปเดต Excel เป็นขั้นตอนสุดท้าย
        if target_sheet == "PendingCust":
            # ย้ายแถวไปยัง Cust หลัก
            ws.cell(row=r, column=5).value = "ยืนยัน"
            ws.cell(row=r, column=19).value = "ชำระแล้ว"
            ws.cell(row=r, column=22).value = f"ยืนยันคิวและล็อก GCal สำเร็จ (Event ID: {gcal_event_id})"
            
            # ย้ายข้อมูลและลบแถว
            move_row_to_confirmed_sheet(wb, ws, r)
        else: # Bookings
            col_gcal, col_status = resolve_bookings_columns(ws)
            ws.cell(row=r, column=col_gcal).value = gcal_event_id
            ws.cell(row=r, column=col_status).value = "ชำระแล้ว"

        save_success = save_workbook_safe(wb, EXCEL_PATH)
        if not save_success:
            return {"success": False, "message": "❌ บันทึกลง Excel ไม่สำเร็จ เนื่องจากไฟล์ถูกเปิดใช้งานอยู่"}
            
        # 5. ส่ง LINE แจ้งลูกค้า
        if user_id:
            booking_date_formatted = format_date_to_dd_mm_yyyy(str(booking_date))
            reply_text = (
                f"✨ ยืนยันการจองคิวสำเร็จแล้วค่ะ คุณ {customer_name}! 🙏✨\n\n"
                f"ทางทีมงานได้ตรวจสอบสลิปการโอนเงินเรียบร้อยแล้วค่ะ และได้ลงคิวนัดหมายของคุณเรียบร้อยแล้วนะคะ 📝\n\n"
                f"📋 รายละเอียดคิวของท่าน:\n"
                f"• เลขอ้างอิง: {queue_id_clean}\n"
                f"• แพคเกจ: {package}\n"
                f"• หมอดู: {astrologer}\n"
                f"• วันนัดหมาย: {booking_date_formatted}\n"
                f"• ช่วงเวลานัด: {booking_slot}\n\n"
                f"📞 เมื่อถึงเวลานัดหมาย จะมีการติดต่อกลับโดยการโทรทาง LINE OA นะคะ\n\n"
                f"ขอให้มีช่วงเวลาทำนายดวงที่ดีนะคะ! ขอบคุณค่ะ 💖"
            )
            push_line_message(user_id, reply_text)
            customer_notified = True
        else:
            customer_notified = False
            
        return {
            "success": True, 
            "message": f"✅ ยืนยันคิว {queue_id_clean} สำเร็จ (ลงตารางงาน และบันทึกลง Excel เรียบร้อยแล้ว{'พร้อมแจ้งลูกค้าแล้ว' if customer_notified else ''})"
        }
        
    except Exception as e:
        print(f"[Admin Command] Error: {e}")
        return {"success": False, "message": f"❌ เกิดข้อผิดพลาดของระบบ: {str(e)}"}

def process_excel_auto_confirmations():
    """
    สแกนตาราง Excel เพื่อตรวจหาคิวที่แอดมินแก้ไขสถานะเป็น 'ยืนยัน' หรือ 'ชำระแล้ว'
    ในชีต PendingCust (และชีต Bookings) จากนั้นทำการจอง Google Calendar และย้ายเข้าชีตหลัก พร้อมแจ้งเตือนลูกค้า
    """
    if not os.path.exists(EXCEL_PATH):
        return
        
    try:
        wb = load_workbook(EXCEL_PATH)
        changes_made = False
        
        # 1. ตรวจสอบชีต PendingCust (คิวจากแชทรอยืนยัน)
        if "PendingCust" in wb.sheetnames:
            ws = wb["PendingCust"]
            # วนลูปถอยหลังเพื่อความปลอดภัยเวลาลบแถว (delete_rows)
            for r in range(ws.max_row, 2, -1):
                status = ws.cell(row=r, column=5).value          # Column E
                pay_status = ws.cell(row=r, column=19).value      # Column S
                
                is_confirmed = (status == "ยืนยัน") or (pay_status == "ชำระแล้ว")
                
                if is_confirmed:
                    customer_name = ws.cell(row=r, column=7).value or "ลูกค้า"
                    line_id = ws.cell(row=r, column=8).value or ""
                    user_id = ws.cell(row=r, column=4).value or ""
                    package = ws.cell(row=r, column=13).value or ""
                    astrologer = ws.cell(row=r, column=14).value or ""
                    booking_date = ws.cell(row=r, column=15).value or ""
                    booking_slot = ws.cell(row=r, column=16).value or ""
                    questions = ws.cell(row=r, column=17).value or ""
                    queue_id = ws.cell(row=r, column=18).value or "Q-xxxx"
                    
                    print(f"[Monitor] Auto-confirming PendingCust row {r} (Queue: {queue_id})")
                    
                    cal_result = check_and_block_calendar_event(
                        date_str=str(booking_date),
                        slot_str=str(booking_slot),
                        name=str(customer_name),
                        line_id=str(line_id),
                        service=str(package),
                        astrologer=str(astrologer),
                        questions=str(questions)
                    )
                    
                    if cal_result.get("success"):
                        gcal_event_id = cal_result.get("eventId")
                        ws.cell(row=r, column=5).value = "ยืนยัน"
                        ws.cell(row=r, column=19).value = "ชำระแล้ว"
                        ws.cell(row=r, column=22).value = f"ยืนยันคิวและล็อก GCal สำเร็จ (Event ID: {gcal_event_id})"
                        
                        # ย้ายข้อมูลไปยัง Cust หลัก และลบแถวออกจาก PendingCust
                        move_row_to_confirmed_sheet(wb, ws, r)
                        changes_made = True
                        
                        if user_id:
                            booking_date_formatted = format_date_to_dd_mm_yyyy(str(booking_date))
                            reply_text = (
                                f"✨ ยืนยันการจองคิวสำเร็จแล้วค่ะ คุณ {customer_name}! 🙏✨\n\n"
                                f"ทางทีมงานได้ตรวจสอบสลิปการโอนเงินเรียบร้อยแล้วค่ะ และได้ลงคิวนัดหมายของคุณเรียบร้อยแล้วนะคะ 📝\n\n"
                                f"📋 รายละเอียดคิวของท่าน:\n"
                                f"• เลขอ้างอิง: {queue_id}\n"
                                f"• แพคเกจ: {package}\n"
                                f"• หมอดู: {astrologer}\n"
                                f"• วันนัดหมาย: {booking_date_formatted}\n"
                                f"• ช่วงเวลานัด: {booking_slot}\n\n"
                                f"📞 เมื่อถึงเวลานัดหมาย จะมีการติดต่อกลับโดยการโทรทาง LINE OA นะคะ\n\n"
                                f"ขอให้มีช่วงเวลาทำนายดวงที่ดีนะคะ! ขอบคุณค่ะ 💖"
                            )
                            push_line_message(user_id, reply_text)
                    else:
                        print(f"[Monitor] Calendar check failed for row {r}: {cal_result.get('message')}")
                        
        # 2. ตรวจสอบชีต Bookings (คิวจากหน้าเว็บ)
        if "Bookings" in wb.sheetnames:
            ws = wb["Bookings"]
            col_gcal, col_status = resolve_bookings_columns(ws)
            for r in range(2, ws.max_row + 1):
                status = ws.cell(row=r, column=col_status).value  # หา column จาก header "Status"
                gcal_id = ws.cell(row=r, column=col_gcal).value   # หา column จาก header "GCal Event ID"

                is_paid = (status == "ชำระแล้ว")
                has_gcal = gcal_id is not None and str(gcal_id).strip() != ""
                
                if is_paid and not has_gcal:
                    queue_id = ws.cell(row=r, column=2).value or "Q-xxxx"
                    customer_name = ws.cell(row=r, column=3).value or "ลูกค้า"
                    line_id = ws.cell(row=r, column=4).value or ""
                    user_id = find_user_id_by_line_id(line_id) or ""
                    package = ws.cell(row=r, column=6).value or ""
                    astrologer = ws.cell(row=r, column=5).value or ""
                    booking_date = ws.cell(row=r, column=7).value or ""
                    booking_slot = ws.cell(row=r, column=8).value or ""
                    questions = ws.cell(row=r, column=9).value or ""
                    
                    print(f"[Monitor] Auto-confirming Bookings row {r} (Queue: {queue_id})")
                    
                    cal_result = check_and_block_calendar_event(
                        date_str=str(booking_date),
                        slot_str=str(booking_slot),
                        name=str(customer_name),
                        line_id=str(line_id),
                        service=str(package),
                        astrologer=str(astrologer),
                        questions=str(questions)
                    )
                    
                    if cal_result.get("success"):
                        gcal_event_id = cal_result.get("eventId")
                        ws.cell(row=r, column=col_gcal).value = gcal_event_id
                        changes_made = True
                        
                        if user_id:
                            booking_date_formatted = format_date_to_dd_mm_yyyy(str(booking_date))
                            reply_text = (
                                f"✨ ยืนยันการจองคิวสำเร็จแล้วค่ะ คุณ {customer_name}! 🙏✨\n\n"
                                f"ทางทีมงานได้ตรวจสอบสลิปการโอนเงินเรียบร้อยแล้วค่ะ และได้ลงคิวนัดหมายของคุณเรียบร้อยแล้วนะคะ 📝\n\n"
                                f"📋 รายละเอียดคิวของท่าน:\n"
                                f"• เลขอ้างอิง: {queue_id}\n"
                                f"• แพคเกจ: {package}\n"
                                f"• หมอดู: {astrologer}\n"
                                f"• วันนัดหมาย: {booking_date_formatted}\n"
                                f"• ช่วงเวลานัด: {booking_slot}\n\n"
                                f"📞 เมื่อถึงเวลานัดหมาย จะมีการติดต่อกลับโดยการโทรทาง LINE OA นะคะ\n\n"
                                f"ขอให้มีช่วงเวลาทำนายดวงที่ดีนะคะ! ขอบคุณค่ะ 💖"
                            )
                            push_line_message(user_id, reply_text)
                    else:
                        print(f"[Monitor] Calendar check failed for Bookings row {r}: {cal_result.get('message')}")
                        
        if changes_made:
            save_workbook_safe(wb, EXCEL_PATH)
            print("[Monitor] Excel database updated with auto-confirmed calendar entries.")
            
    except Exception as e:
        print(f"[Monitor] Error scanning Excel database: {e}")

def excel_monitor_thread_func():
    """
    ฟังก์ชันตรวจสอบการอัปเดตไฟล์ Excel แบบเบื้องหลัง (Background Monitor)
    ตรวจจับเมื่อผู้ดูแลระบบเซฟไฟล์ Excel แล้วประมวลผลคิวที่มีสถานะเป็น 'ยืนยัน'
    """
    print("[Monitor] Excel Monitor Thread started.")
    last_mtime = 0
    if os.path.exists(EXCEL_PATH):
        try:
            last_mtime = os.path.getmtime(EXCEL_PATH)
        except Exception:
            last_mtime = 0
            
    while True:
        try:
            time.sleep(10)  # ตรวจสอบทุกๆ 10 วินาที
            if not os.path.exists(EXCEL_PATH):
                continue
                
            try:
                current_mtime = os.path.getmtime(EXCEL_PATH)
            except Exception:
                continue
                
            if current_mtime != last_mtime:
                print(f"[Monitor] Excel file modification detected.")
                time.sleep(3) # รอให้การแก้ไขปล่อยไฟล์เสร็จ
                process_excel_auto_confirmations()
                
                try:
                    last_mtime = os.path.getmtime(EXCEL_PATH)
                except Exception:
                    pass
        except Exception as e:
            print(f"[Monitor] Error in monitor loop: {e}")

def download_line_image(message_id):
    """ดาวน์โหลดไฟล์รูปภาพสลิปจาก LINE Messaging API"""
    if not LINE_CHANNEL_ACCESS_TOKEN:
        print("[LINE Image] Skipping: LINE_CHANNEL_ACCESS_TOKEN not configured")
        return None
    url = f"https://api-data.line.me/v2/bot/message/{message_id}/content"
    headers = {
        "Authorization": f"Bearer {LINE_CHANNEL_ACCESS_TOKEN}"
    }
    try:
        res = requests.get(url, headers=headers, timeout=15)
        if res.status_code == 200:
            return res.content
        else:
            print(f"[LINE Image] Failed to download {message_id}: {res.status_code} {res.text}")
            return None
    except Exception as e:
        print(f"[LINE Image] Error fetching image {message_id}: {e}")
        return None

def verify_slip_via_slipok(image_bytes, expected_amount=None):
    """
    ตรวจสอบความถูกต้องของสลิปโอนเงินผ่าน SlipOK API
    """
    if not SLIPOK_API_KEY or not SLIPOK_BRANCH_ID:
        print("[SlipOK] Skipping: API Key or Branch ID not configured")
        return {"success": False, "error": "not_configured", "message": "ไม่ได้ตั้งค่า SlipOK API/Branch ID"}
        
    url = f"https://api.slipok.com/api/line/apikey/{SLIPOK_BRANCH_ID}"
    headers = {
        "x-authorization": SLIPOK_API_KEY
    }
    
    files = {
        "files": ("slip.png", image_bytes, "image/png")
    }
    
    data = {
        "log": "true"
    }
    if expected_amount:
        data["amount"] = expected_amount
        
    try:
        res = requests.post(url, headers=headers, files=files, data=data, timeout=15)
        if res.status_code == 200:
            # ผลลัพธ์ของ SlipOK จะอยู่ในรูปแบบ JSON
            return res.json()
        else:
            print(f"[SlipOK] Error HTTP {res.status_code}: {res.text}")
            return {"success": False, "error": "http_error", "message": f"HTTP {res.status_code} {res.text[:100]}"}
    except Exception as e:
        print(f"[SlipOK] Exception checking slip: {e}")
        return {"success": False, "error": "exception", "message": str(e)}

def find_pending_booking_by_user_id(user_id):
    """
    ค้นหารายการจองที่รอยืนยันในชีต PendingCust จาก LINE User ID (Column D)
    """
    if not os.path.exists(EXCEL_PATH) or not user_id:
        return None
    try:
        wb = load_workbook(EXCEL_PATH)
        if "PendingCust" not in wb.sheetnames:
            return None
        ws = wb["PendingCust"]
        for r in range(3, ws.max_row + 1):
            val = ws.cell(row=r, column=4).value  # Column D: LINE User ID
            if val and str(val).strip() == str(user_id).strip():
                # ส่งกลับข้อมูลเพื่อนำไปใช้ต่อ
                return {
                    "row": r,
                    "queue_id": ws.cell(row=r, column=18).value or "",
                    "customer_name": ws.cell(row=r, column=7).value or "ลูกค้า",
                    "package": ws.cell(row=r, column=13).value or "",
                    "astrologer": ws.cell(row=r, column=14).value or "",
                    "booking_date": ws.cell(row=r, column=15).value or "",
                    "booking_slot": ws.cell(row=r, column=16).value or "",
                }
        return None
    except Exception as e:
        print(f"[Excel] Error in find_pending_booking_by_user_id: {e}")
        return None

def check_calendar_slot_busy(date_str, slot_str):
    """
    ตรวจสอบว่าช่วงเวลาใน Google Calendar มีการจองชนกันหรือไม่ (ตรวจสอบอย่างเดียว ไม่จอง)
    """
    if not GOOGLE_SCRIPT_URL:
        print("[Calendar] Skipping: GOOGLE_SCRIPT_URL not configured")
        return {"success": True, "busy": False}
        
    try:
        start_iso, end_iso = parse_date_slot_to_iso(date_str, slot_str)
    except Exception as e:
        print(f"[Calendar] Date/Time parsing failed: {e}")
        return {"success": False, "error": "parse_error", "message": str(e)}
        
    payload = {
        "action": "checkSlot",
        "startTime": start_iso,
        "endTime": end_iso
    }
    
    try:
        res = requests.post(GOOGLE_SCRIPT_URL, json=payload, timeout=10)
        if res.status_code == 200:
            return res.json()
        else:
            print(f"[Calendar] Google Script HTTP error ({res.status_code}): {res.text}")
            return {"success": False, "error": "http_error", "message": f"HTTP {res.status_code}"}
    except Exception as e:
        print(f"[Calendar] Request to Google Script failed: {e}")
        return {"success": True, "busy": False, "warning": "connection_error"}

def check_and_block_calendar_event(date_str, slot_str, name, line_id, service, astrologer, questions=""):
    """
    ตรวจสอบและบล็อกสลอตเวลาใน Google Calendar ผ่าน Google Apps Script Web App
    """
    if not GOOGLE_SCRIPT_URL:
        print("[Calendar] Skipping: GOOGLE_SCRIPT_URL not configured")
        return {"success": True, "simulated": True, "message": "ข้ามการตรวจสอบปฏิทิน (ไม่ได้ตั้งค่าลิงก์)"}
        
    try:
        start_iso, end_iso = parse_date_slot_to_iso(date_str, slot_str)
    except Exception as e:
        print(f"[Calendar] Date/Time parsing failed: {e}")
        return {"success": False, "error": "parse_error", "message": str(e)}
        
    summary = f"MuHub: {name} ({astrologer})"
    description = (
        f"คิวจอง: {summary}\n"
        f"แพคเกจ: {service}\n"
        f"ไลน์ลูกค้า: {line_id}\n"
        f"คำถามพิเศษ: {questions}\n"
        f"จองผ่านแชท LINE OA"
    )
    
    payload = {
        "action": "checkAndCreate",
        "startTime": start_iso,
        "endTime": end_iso,
        "summary": summary,
        "description": description
    }
    
    try:
        res = requests.post(GOOGLE_SCRIPT_URL, json=payload, timeout=10)
        if res.status_code == 200:
            return res.json()
        else:
            print(f"[Calendar] Google Script HTTP error ({res.status_code}): {res.text}")
            return {"success": False, "error": "http_error", "message": f"HTTP {res.status_code}"}
    except Exception as e:
        print(f"[Calendar] Request to Google Script failed: {e}")
        return {"success": False, "error": "connection_error", "message": str(e)}

# ── ฟังก์ชันตัวถอดข้อความระบบจอง (Booking Message Parser) ──────────────────────
def parse_booking_message(text):
    """วิเคราะห์แยกแยะฟิลด์จากรูปแบบข้อความที่ลูกค้าส่งเข้า LINE OA จากระบบจองคิว"""
    fields = {
        "name": "", "line_id": "", "birth_date": "", "birth_time": "",
        "country": "", "city": "", "questions": "", "queue_id": "",
        "package": "", "astrologer": "", "booking_date": "", "booking_slot": "",
        "payment_status": ""
    }
    
    lines = text.split("\n")
    for line in lines:
        line = line.strip()
        if not line.startswith("•"):
            continue
            
        parts = line.split(":", 1)
        if len(parts) < 2:
            continue
            
        key = parts[0].replace("•", "").strip()
        val = parts[1].strip()
        
        if "ชื่อ" in key:
            fields["name"] = val
        elif "Line ID" in key or "ไลน์" in key:
            fields["line_id"] = val
        elif "วันเกิด" in key:
            fields["birth_date"] = val
        elif "เวลาเกิด" in key:
            fields["birth_time"] = val
        elif "ประเทศ" in key:
            fields["country"] = val
        elif "จังหวัด" in key:
            fields["city"] = val
        elif "คำถามพิเศษ" in key:
            fields["questions"] = val
        elif "เลขอ้างอิง" in key:
            fields["queue_id"] = val
        elif "แพคเกจ" in key:
            fields["package"] = val
        elif "หมอดู" in key:
            fields["astrologer"] = val
        elif "วันนัด" in key:
            fields["booking_date"] = val
        elif "เวลานัด" in key:
            fields["booking_slot"] = val
        elif "สถานะการชำระเงิน" in key:
            fields["payment_status"] = val
            
    return fields

# ── ฟังก์ชันการใช้งาน LINE APIs ───────────────────────────────────────────
def find_user_id_by_line_id(line_id):
    """ค้นหา LINE User ID (รหัสขึ้นต้นด้วย U...) จาก Line ID (username) ในชีต Cust"""
    if not line_id:
        return None
    line_id_clean = str(line_id).strip().lower()
    if line_id_clean.startswith('@'):
        line_id_clean = line_id_clean[1:]
        
    if not os.path.exists(EXCEL_PATH):
        return None
        
    try:
        wb = load_workbook(EXCEL_PATH, read_only=True)
        if "Cust" in wb.sheetnames:
            ws = wb["Cust"]
            for r in range(ws.max_row, 2, -1):
                u_id = ws.cell(row=r, column=4).value # Column D (LINE User ID)
                l_id = ws.cell(row=r, column=8).value # Column H (Line ID username)
                if l_id:
                    l_id_clean = str(l_id).strip().lower()
                    if l_id_clean.startswith('@'):
                        l_id_clean = l_id_clean[1:]
                    if l_id_clean == line_id_clean:
                        if u_id and str(u_id).strip().startswith("U"):
                            return str(u_id).strip()
    except Exception as e:
        print(f"[Lookup] Error finding user_id for line_id '{line_id}': {e}")
    return None

def push_line_message(user_id, text):
    """ส่งข้อความแบบ Push หาผู้ใช้คนใดคนหนึ่งโดยตรงผ่าน LINE OA"""
    if not LINE_CHANNEL_ACCESS_TOKEN or not user_id:
        print(f"[LINE] Skipping push (token or user_id missing)")
        return False
        
    try:
        url = "https://api.line.me/v2/bot/message/push"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {LINE_CHANNEL_ACCESS_TOKEN}"
        }
        payload = {
            "to": user_id,
            "messages": [
                {
                    "type": "text",
                    "text": text
                }
            ]
        }
        res = requests.post(url, headers=headers, json=payload, timeout=5)
        if res.status_code == 200:
            print(f"[LINE] Sent push successfully to {user_id}")
            return True
        else:
            print(f"[LINE] Failed to send push ({res.status_code}): {res.text}")
    except Exception as e:
        print(f"[LINE] Error sending push: {e}")
    return False

def get_line_profile(user_id):
    """ดึงโปรไฟล์ลูกค้าผ่าน LINE API ด้วย caching ป้องกันการยิงซ้ำซ้อน"""
    if user_id in PROFILE_CACHE:
        return PROFILE_CACHE[user_id]
        
    if not LINE_CHANNEL_ACCESS_TOKEN:
        return {"displayName": "ลูกค้า LINE"}
        
    try:
        url = f"https://api.line.me/v2/bot/profile/{user_id}"
        headers = {"Authorization": f"Bearer {LINE_CHANNEL_ACCESS_TOKEN}"}
        res = requests.get(url, headers=headers, timeout=5)
        if res.status_code == 200:
            profile = res.json()
            PROFILE_CACHE[user_id] = profile
            return profile
        else:
            print(f"[LINE] Failed to get profile ({res.status_code}): {res.text}")
    except Exception as e:
        print(f"[LINE] Error getting profile: {e}")
        
    return {"displayName": "ลูกค้า LINE"}

def reply_line_message(reply_token, reply_text):
    """ส่งข้อความตอบกลับหาลูกค้าผ่านระบบ LINE Messaging API"""
    if not LINE_CHANNEL_ACCESS_TOKEN or not reply_token:
        return
        
    try:
        url = "https://api.line.me/v2/bot/message/reply"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {LINE_CHANNEL_ACCESS_TOKEN}"
        }
        payload = {
            "replyToken": reply_token,
            "messages": [
                {
                    "type": "text",
                    "text": reply_text
                }
            ]
        }
        res = requests.post(url, headers=headers, json=payload, timeout=5)
        if res.status_code == 200:
            print(f"[LINE] Sent reply successfully")
        else:
            print(f"[LINE] Failed to send reply ({res.status_code}): {res.text}")
    except Exception as e:
        print(f"[LINE] Error sending reply: {e}")

def verify_line_signature(request_data, signature):
    """ตรวจสอบลายเซ็นจาก LINE Webhook ป้องกันการยิงสแปมแกล้ง"""
    if not LINE_CHANNEL_SECRET:
        # ข้ามการตรวจสอบหากผู้ใช้ยังไม่ได้ตั้งค่า Secret
        return True
        
    hash = hmac.new(
        LINE_CHANNEL_SECRET.encode('utf-8'),
        request_data,
        hashlib.sha256
    ).digest()
    calculated_signature = base64.b64encode(hash).decode('utf-8')
    return hmac.compare_digest(calculated_signature, signature)

# ── API ROUTES ───────────────────────────────────────────────────────────
@app.route("/api/save-customer", methods=["POST", "OPTIONS"])
def save_customer():
    """รองรับฟังก์ชันจองคิวเดิมจากหน้าเว็บ โดยบันทึกลงชีต Bookings หลัก"""
    if request.method == "OPTIONS":
        return "", 200

    guard_result = require_app_token_and_rate_limit("save_customer")
    if guard_result is not None:
        return guard_result

    try:
        data = request.get_json(force=True)
        print(f"[Direct Web API] Received booking: {data.get('name')} / {data.get('queueId')}")

        wb = get_or_create_workbook()
        # ทำการเขียนเข้าชีต Bookings หลักตามเซิร์ฟเวอร์เดิม
        if "Bookings" in wb.sheetnames:
            ws = wb["Bookings"]
        else:
            ws = wb.active

        row = [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            sanitize_excel_value(data.get("queueId",      "")),
            sanitize_excel_value(data.get("name",         "")),
            sanitize_excel_value(data.get("lineId",       "")),
            sanitize_excel_value(data.get("astrologer",   "")),
            sanitize_excel_value(data.get("service",      "")),
            sanitize_excel_value(data.get("date",         "")),
            sanitize_excel_value(data.get("slot",         "")),
            sanitize_excel_value(data.get("questions",    "")),
            sanitize_excel_value(data.get("birthDate",    "")),
            sanitize_excel_value(data.get("birthHour",    "")),
            sanitize_excel_value(data.get("birthMin",     "")),
            sanitize_excel_value(data.get("birthCountry", "")),
            sanitize_excel_value(data.get("birthCity",    "")),
            sanitize_excel_value(data.get("gcalEventId",  "")),
            "pending"
        ]

        ws.append(row)

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

        success = save_workbook_safe(wb, EXCEL_PATH)
        if success:
            print(f"[Direct Web API] Saved directly to {EXCEL_PATH} (row {ws.max_row})")
            
            # ดึง LINE ID และ ข้อความสำหรับส่งไลน์
            line_id = data.get("lineId", "")
            line_message = data.get("lineMessage", "")
            
            line_sent = False
            
            # ค้นหา user_id จาก line_id (username) ในชีต Cust
            user_id = find_user_id_by_line_id(line_id)
            if user_id:
                print(f"[Direct Web API] Found user_id '{user_id}' for LINE ID '{line_id}'")
                if line_message:
                    line_sent = push_line_message(user_id, line_message)
            else:
                print(f"[Direct Web API] No user_id found for LINE ID '{line_id}' in Cust database.")
                
            # ส่งแจ้งเตือน Admin (ถ้ามีการตั้งค่าไว้)
            if LINE_ADMIN_USER_ID and line_message:
                admin_text = f"🔔 [ระบบจองคิว] มีคิวใหม่จองเข้ามาค่ะ!\n\n{line_message}"
                push_line_message(LINE_ADMIN_USER_ID, admin_text)
                
            return jsonify({"success": True, "row": ws.max_row, "queueId": data.get("queueId"), "lineSent": line_sent})
        else:
            return jsonify({"success": False, "error": "Excel file is locked"}), 500

    except Exception as e:
        print(f"[Direct Web API] Error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/line-webhook", methods=["POST"])
@app.route("/webhook", methods=["POST"])
def line_webhook():
    """ตัวรับ Webhook จาก LINE OA บันทึกฐานข้อมูลอัตโนมัติลงชีต Cust"""
    signature = request.headers.get("X-Line-Signature", "")
    body_bytes = request.get_data()
    
    if not verify_line_signature(body_bytes, signature):
        print("[Webhook] Warning: Invalid signature, request rejected.")
        return "Invalid signature", 400
        
    try:
        payload = request.get_json(force=True)
    except Exception as e:
        print(f"[Webhook] Error decoding JSON body: {e}")
        return "Bad request", 400
        
    events = payload.get("events", [])
    print(f"[Webhook] Received webhook event from LINE. Count: {len(events)}")
    
    for event in events:
        event_type = event.get("type")
        source = event.get("source", {})
        source_type = source.get("type")
        
        if event_type != "message" or source_type != "user":
            continue
            
        message = event.get("message", {})
        message_type = message.get("type")
        
        if message_type not in ["text", "image"]:
            continue
            
        user_id = source.get("userId")
        reply_token = event.get("replyToken")
        
        # ดึงชื่อเล่นหรือโปรไฟล์จาก LINE
        profile = get_line_profile(user_id)
        display_name = profile.get("displayName", "ลูกค้า LINE")
        
        if message_type == "image":
            message_id = message.get("id")
            print(f"[Webhook] Image message received from {user_id} (Message ID: {message_id})")
            
            # ค้นหารายการจองใน PendingCust
            pending_booking = find_pending_booking_by_user_id(user_id)
            if not pending_booking:
                # ถ้าไม่พบรายการจอง ไม่ต้องประมวลผลเป็นสลิป
                print(f"[Webhook] Image received from {user_id} but no pending booking found.")
                continue
                
            queue_id = pending_booking["queue_id"]
            customer_name = pending_booking["customer_name"]
            
            if not SLIPOK_API_KEY or not SLIPOK_BRANCH_ID:
                # ไม่ได้เปิดใช้ API ให้แจ้งเตือนแอดมินให้มาตรวจสลิปเองแบบแมนนวล
                print(f"[Webhook] SlipOK is not configured. Notify admin for manual review.")
                if LINE_ADMIN_USER_ID:
                    admin_notify = (
                        f"🔔 [แจ้งโอนเงิน] คุณ {customer_name} ส่งรูปสลิปเข้ามาแล้วสำหรับคิว {queue_id} ค่ะ\n"
                        f"กรุณาตรวจสอบสลิปแล้วพิมพ์:\n"
                        f"👉 ยืนยัน {queue_id}\n"
                        f"หรือแก้ไขแถวใน Excel เป็น 'ยืนยัน'"
                    )
                    push_line_message(LINE_ADMIN_USER_ID, admin_notify)
                
                reply_line_message(
                    reply_token,
                    f"🙏 ได้รับหลักฐานการโอนเงินเรียบร้อยแล้วค่ะ คุณ {customer_name}!\n\n"
                    f"แอดมินจะรีบทำการตรวจสอบสลิปและแจ้งเตือนยืนยันคิวให้ท่านโดยเร็วที่สุดนะคะ 📝"
                )
                continue
                
            # มี SlipOK -> ดาวน์โหลดและตรวจสอบสลิป
            image_bytes = download_line_image(message_id)
            if not image_bytes:
                reply_line_message(
                    reply_token,
                    "❌ ไม่สามารถดาวน์โหลดรูปภาพสลิปได้ชั่วคราว รบกวนลองส่งสลิปใหม่อีกครั้ง หรือรอแอดมินตรวจสอบนะคะ 🙏"
                )
                continue
                
            print(f"[Webhook] Verifying slip via SlipOK for Queue: {queue_id}...")
            res_slip = verify_slip_via_slipok(image_bytes)
            
            if res_slip.get("success") and res_slip.get("data", {}).get("success"):
                slip_data = res_slip["data"]
                amount = slip_data.get("amount", 0)
                trans_ref = slip_data.get("transRef", "")
                
                print(f"[Webhook] SlipOK verified successfully: {amount} THB (Ref: {trans_ref})")
                
                # ทำการอนุมัติล็อกปฏิทินและบันทึก Cust อัตโนมัติ!
                confirm_res = confirm_and_gcal_booking(queue_id)
                
                if confirm_res.get("success"):
                    # ส่งแจ้งเตือนแอดมิน
                    if LINE_ADMIN_USER_ID:
                        admin_notify = (
                            f"✅ [ตรวจสอบสลิปอัตโนมัติ] ระบบตรวจผ่านสลิปของ คุณ {customer_name} (คิว: {queue_id}) ผ่าน SlipOK เรียบร้อยแล้วค่ะ\n"
                            f"• ยอดโอน: {amount} บาท\n"
                            f"• รหัสอ้างอิง: {trans_ref}\n"
                            f"• ลงปฏิทินและบันทึกลงฐานข้อมูลหลัก Cust สำเร็จแล้วค่ะ"
                        )
                        push_line_message(LINE_ADMIN_USER_ID, admin_notify)
                else:
                    err_msg = confirm_res.get("message", "ไม่สามารถบันทึกได้")
                    reply_line_message(
                        reply_token,
                        f"⚠️ ตรวจสอบสลิปผ่านแล้ว แต่ระบบพบล็อกเวลาไม่ว่างหรือขัดข้อง: {err_msg}\n"
                        f"ไม่ต้องเป็นกังวลนะคะ แอดมินกำลังเร่งดำเนินการล็อกคิวให้ท่านแบบแมนนวลค่ะ 🙏"
                    )
                    if LINE_ADMIN_USER_ID:
                        push_line_message(
                            LINE_ADMIN_USER_ID,
                            f"⚠️ [ตรวจสอบสลิปอัตโนมัติ] ตรวจผ่านสลิป ({amount} บาท) แต่จองลงปฏิทิน/บันทึก Excel ล้มเหลวสำหรับคิว {queue_id}: {err_msg}"
                        )
            else:
                # ตรวจสอบสลิปไม่สำเร็จ (สลิปปลอม, สลิปเก่า, สลิปยอดไม่ตรง, หรือไม่มี QR)
                err_msg = res_slip.get("message") or "ไม่พบ QR Code บนสลิป หรือเป็นสลิปซ้ำ"
                print(f"[Webhook] SlipOK verification failed: {err_msg}")
                
                reply_line_message(
                    reply_token,
                    f"❌ ระบบไม่สามารถตรวจสอบข้อมูลบนรูปภาพสลิปได้โดยอัตโนมัติ ({err_msg}) 😢\n\n"
                    f"แต่ไม่ต้องกังวลใจไปนะคะ! แอดมินจะทำการตรวจสอบสลิปการโอนเงินของท่านด้วยตนเองในแชทนี้ให้โดยเร็วที่สุดค่ะ 🙏"
                )
                
                if LINE_ADMIN_USER_ID:
                    push_line_message(
                        LINE_ADMIN_USER_ID,
                        f"⚠️ [ตรวจสอบสลิปอัตโนมัติล้มเหลว] สลิปของ คุณ {customer_name} ({queue_id}) ตรวจไม่ผ่าน: {err_msg}\n"
                        f"รบกวนแอดมินเข้ามาตรวจสอบรูปภาพสลิปด้วยตนเองนะคะ"
                    )
            continue
            
        # ถ้าเป็น text
        message_text = message.get("text", "")
        
        # ตรวจสอบว่าเป็นคำสั่งจากแอดมินในการยืนยันคิวหรือไม่
        is_admin_cmd = (user_id == LINE_ADMIN_USER_ID) and (
            message_text.strip().startswith("ยืนยัน ") or 
            message_text.strip().lower().startswith("/confirm ")
        )
        
        if is_admin_cmd:
            print(f"[Webhook] Admin command matched: {message_text}")
            parts = message_text.strip().split(" ", 1)
            queue_id_cmd = parts[1].strip() if len(parts) > 1 else ""
            
            result = confirm_and_gcal_booking(queue_id_cmd)
            reply_line_message(reply_token, result["message"])
            return "OK", 200
            
        # ตรวจสอบว่าเป็นข้อความแพทเทิร์นการจองหรือไม่
        is_booking = "📋 ข้อมูลวันเกิดสำหรับทำนายดวง" in message_text
        
        if is_booking:
            print(f"[Webhook] Matching booking message from customer {display_name}")
            booking_data = parse_booking_message(message_text)
            
            # ตรวจสอบความว่างของเวลานัดหมายในปฏิทินก่อน
            cal_check = check_calendar_slot_busy(
                date_str=booking_data.get("booking_date", ""),
                slot_str=booking_data.get("booking_slot", "")
            )
            
            if not cal_check.get("success") and cal_check.get("error") == "parse_error":
                # รูปแบบวันที่/เวลา ผิดพลาด
                reply_text = (
                    f"ขออภัยค่ะ รูปแบบวันที่หรือเวลานัดที่ระบุไม่ถูกต้อง: {cal_check.get('message')} 😢\n\n"
                    f"กรุณากรอกข้อมูลวันนัดและเวลานัดให้ถูกต้องตามรูปแบบตัวอย่าง (เช่น 2026-06-30 และ 10:00 - 11:00) และส่งแบบฟอร์มอีกครั้งนะคะ 🙏"
                )
                reply_line_message(reply_token, reply_text)
                return "OK", 200
                
            if cal_check.get("busy"):
                # กรณีคิวไม่ว่าง แจ้งเตือนลูกค้าทันที
                reply_text = (
                    f"ขออภัยค่ะ คุณ {booking_data.get('name')}! 🙏✨\n"
                    f"วันนัดหมายและเวลาที่คุณเลือก ({booking_data.get('booking_date')} ช่วง {booking_data.get('booking_slot')}) "
                    f"**มีผู้อื่นจองคิวไว้เรียบร้อยแล้วค่ะ** 😢\n\n"
                    f"กรุณาเลือกวันนัดหมายหรือช่วงเวลาอื่น และส่งแบบฟอร์มจองเข้ามาใหม่อีกครั้งนะคะ ขออภัยในความไม่สะดวกค่ะ 💖"
                )
                reply_line_message(reply_token, reply_text)
                return "OK", 200
                
            data = {
                "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M"),
                "display_name": display_name,
                "user_id": user_id,
                "status": "ใหม่",
                "booking_date": datetime.now().strftime("%d/%m/%Y"),
                "customer_name": booking_data.get("name", ""),
                "line_id": booking_data.get("line_id", ""),
                "birth_date": booking_data.get("birth_date", ""),
                "birth_time": booking_data.get("birth_time", ""),
                "birth_country": booking_data.get("country", ""),
                "birth_city": booking_data.get("city", ""),
                "package": booking_data.get("package", ""),
                "astrologer": booking_data.get("astrologer", ""),
                "booking_date_appointment": booking_data.get("booking_date", ""),
                "booking_slot": booking_data.get("booking_slot", ""),
                "questions": booking_data.get("questions", ""),
                "queue_id": booking_data.get("queue_id", ""),
                "payment_status": "รอตรวจสอบ",
                "notes": "จองผ่านแชท LINE OA (รอตรวจสอบสลิป)"
            }
            
            # บันทึกลง Excel (ในชีต PendingCust ชั่วคราว)
            process_data_write(data, target_sheet="PendingCust")
            
            # ตอบกลับแจ้งลูกค้าเพื่อถามความสะดวกโอนเงินคอนเฟิร์มคิว
            booking_date_formatted = format_date_to_dd_mm_yyyy(booking_data.get('booking_date', ''))
            reply_text = (
                f"ได้รับข้อมูลการจองคิวเรียบร้อยแล้วค่ะ คุณ {booking_data.get('name')}! 🙏✨\n\n"
                f"📋 รายละเอียดการจองคิวของท่าน:\n"
                f"• เลขอ้างอิง: {booking_data.get('queue_id')}\n"
                f"• แพคเกจ: {booking_data.get('package')}\n"
                f"• หมอดู: {booking_data.get('astrologer')}\n"
                f"• วันนัดหมาย: {booking_date_formatted}\n"
                f"• ช่วงเวลานัด: {booking_data.get('booking_slot')}\n\n"
                f"💵 ช่วงเวลานี้ว่างและสามารถจองได้ค่ะ ไม่ทราบว่าคุณสะดวกโอนเงินเพื่อคอนเฟิร์มคิวเลยไหมคะ? (หากโอนเงินเรียบร้อยแล้ว รบกวนส่งรูปภาพสลิปการโอนเงินยืนยันเข้ามาในแชทนี้ได้เลยนะคะ) 💖"
            )
            reply_line_message(reply_token, reply_text)
            return "OK", 200
            
        else:
            # ข้อความทั่วไปทักทายลูกค้า
            print(f"[Webhook] Regular message from customer {display_name}")
            
            # เปิดเช็คว่าลูกค้าอยู่ในฐานข้อมูลอยู่แล้วหรือไม่
            wb_exists = os.path.exists(EXCEL_PATH)
            has_user = False
            if wb_exists:
                try:
                    wb = load_workbook(EXCEL_PATH, read_only=True)
                    if "Cust" in wb.sheetnames:
                        ws = wb["Cust"]
                        for r in range(3, ws.max_row + 1):
                            if ws.cell(row=r, column=4).value == user_id:
                                has_user = True
                                break
                except Exception:
                    has_user = False
                    
            # คำนวณความสนใจในการจองคิวจากคีย์เวิร์ด
            booking_keywords = ["จอง", "ดูดวง", "คิว", "package", "แพคเกจ", "หมอดู", "ผูกดวง", "ทำนาย", "จองคิว", "book", "slot"]
            is_asking_booking = any(kw in message_text.lower() for kw in booking_keywords)
            
            if has_user and not is_asking_booking:
                # กรณีมีประวัติอยู่แล้ว และไม่ได้แชทถามเรื่องจองคิว ให้แอดข้อความลงในช่องหมายเหตุท้ายแถว
                print(f"[Webhook] User exists. Appending chat note to Excel.")
                data = {
                    "notes": message_text
                }
                process_data_write(data, append_only_user_id=user_id)
                # เพื่อความเป็นส่วนตัวและไม่รบกวนการคุยสด จึงไม่ต้องส่งบอทแชทกลับหาลูกค้าเก่า
            else:
                # กรณีไม่เคยทักมาเลย หรือทักมาถามเรื่องจองคิว ให้สร้างประวัติผู้สนใจ (ถ้าเป็นลูกค้าใหม่) และส่งคำแนะนำการจอง
                if not has_user:
                    print(f"[Webhook] New user. Creating new customer row in Excel.")
                    data = {
                        "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M"),
                        "display_name": display_name,
                        "user_id": user_id,
                        "status": "ผู้สนใจใหม่",
                        "notes": f"ทักแชทครั้งแรก: {message_text}"
                    }
                    process_data_write(data)
                
                # ตอบกลับต้อนรับลูกค้าใหม่/ผู้ต้องการจองคิว
                reply_text = (
                    f"สวัสดีค่ะ คุณ {display_name} ยินดีต้อนรับสู่ MuHub ค่ะ! 🙏✨\n\n"
                    f"หากคุณต้องการจองคิวดูดวง/ผูกดวงชะตา สามารถทำรายการได้ 2 ช่องทางดังนี้ค่ะ:\n\n"
                    f"🌟 ช่องทางที่ 1: จองผ่านระบบเว็บแอปพลิเคชัน (แนะนำที่สุด ⚡ สะดวก รวดเร็ว เลือกวันและเวลาได้ทันที)\n"
                    f"👉 สามารถคลิกทำรายการได้ที่: {WEB_APP_URL}\n\n"
                    f"📝 ช่องทางที่ 2: จองผ่านช่องแชทนี้โดยตรง\n"
                    f"กรุณาคัดลอกแบบฟอร์มด้านล่างนี้ กรอกข้อมูลให้ครบถ้วน และส่งกลับมาในแชทนี้พร้อมแนบภาพสลิปโอนเงินนะคะ:\n\n"
                    f"📋 ข้อมูลวันเกิดสำหรับทำนายดวง\n"
                    f"• ชื่อ:\n"
                    f"• Line ID:\n"
                    f"• วันเกิด:\n"
                    f"• เวลาเกิด:\n"
                    f"• ประเทศ: ประเทศไทย\n"
                    f"• จังหวัด:\n"
                    f"• คำถามพิเศษ:\n\n"
                    f"💳 รายละเอียดการจอง\n"
                    f"• เลขอ้างอิง: (ระบุ 'จองผ่านแชท')\n"
                    f"• แพคเกจ:\n"
                    f"• หมอดู:\n"
                    f"• วันนัด:\n"
                    f"• เวลานัด:"
                )
                reply_line_message(reply_token, reply_text)
                
    return "OK", 200

def _load_promo_used():
    """อ่านไฟล์ promo_used.json -> {"MUHUBFIRST": ["lineid1", "lineid2", ...]}"""
    if os.path.exists(PROMO_USED_PATH):
        try:
            with open(PROMO_USED_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"[Promo] Warning: Failed to read promo_used.json: {e}")
    return {}


def _save_promo_used(data):
    with open(PROMO_USED_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


@app.route("/api/check-promo", methods=["GET"])
def check_promo():
    """
    เช็คว่า Line ID นี้ยังมีสิทธิ์ใช้โค้ดโปรโมชั่นนี้อยู่หรือไม่ (read-only, ไม่ล็อกสิทธิ์)
    เรียกตอนลูกค้ากด "ใช้งาน" โค้ด เพื่อโชว์ผลแบบทันที
    """
    code = (request.args.get("code") or "").strip().upper()
    line_id = (request.args.get("lineId") or "").strip().lower()
    if not code or not line_id:
        return jsonify({"eligible": False, "reason": "missing_params"}), 400

    with promo_lock:
        used = _load_promo_used()
        already_used = line_id in used.get(code, [])

    return jsonify({"eligible": not already_used, "code": code})


@app.route("/api/redeem-promo", methods=["POST"])
def redeem_promo():
    """
    ล็อกสิทธิ์โค้ดโปรโมชั่นให้ Line ID นี้แบบถาวร (เขียนไฟล์จริง)
    เรียกตอนลูกค้ากด "ยืนยันการชำระเงิน" ของการจองแบบใช้โค้ดฟรีเท่านั้น
    เช็ค + เขียนอยู่ใน lock เดียวกันเพื่อกันสองคำขอพร้อมกันแย่งสิทธิ์เดียวกัน
    """
    guard_result = require_app_token_and_rate_limit("redeem_promo")
    if guard_result is not None:
        return guard_result

    data = request.get_json(force=True) or {}
    code = (data.get("code") or "").strip().upper()
    line_id = (data.get("lineId") or "").strip().lower()
    if not code or not line_id:
        return jsonify({"success": False, "reason": "missing_params"}), 400

    with promo_lock:
        used = _load_promo_used()
        used_list = used.setdefault(code, [])
        if line_id in used_list:
            return jsonify({"success": False, "reason": "already_used"})
        used_list.append(line_id)
        _save_promo_used(used)

    return jsonify({"success": True})


def _lookup_booking_status(queue_id: str):
    """
    ดึงสถานะการจองจริงจากไฟล์ Excel (แหล่งความจริงเดียว ที่แอดมินแก้ด้วยมือ)
    ใช้เงื่อนไข "ยืนยันแล้ว" แบบเดียวกับที่ process_excel_auto_confirmations()
    ใช้ตัดสินใจส่ง LINE ยืนยันลูกค้าจริง เพื่อให้ endpoint นี้ไม่มีวันบอกว่า
    "คอนเฟิร์ม" ก่อนที่ระบบอัตโนมัติจริงจะยืนยันจริง

    หมายเหตุ: คอลัมน์ "GCal Event ID"/"Status" ของชีต Bookings หาโดยอ่านชื่อ
    header จากแถวที่ 1 ผ่าน resolve_bookings_columns() ก่อนเสมอ (ไม่ hardcode
    เลขคอลัมน์แล้ว) เผื่อไว้ด้วยว่ายังเช็คคอลัมน์ 15/16 ตามที่ HEADERS ประกาศไว้
    เป็นอีกชั้นหนึ่ง กรณีไฟล์จริงไม่มี header ตรงกับที่คาดไว้เลย
    """
    if not os.path.exists(EXCEL_PATH):
        return None
    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
    except Exception as e:
        print(f"[BookingStatus] Failed to open workbook: {e}")
        return None

    # 1. ชีต "Bookings" (คิวที่จองตรงจากหน้าเว็บ)
    if "Bookings" in wb.sheetnames:
        ws = wb["Bookings"]
        col_gcal, col_status = resolve_bookings_columns(ws)
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(row=r, column=2).value or "").strip() == queue_id:
                status_literal = str(ws.cell(row=r, column=16).value or "").strip()
                status_u = str(ws.cell(row=r, column=col_status).value or "").strip()
                gcal_t = str(ws.cell(row=r, column=col_gcal).value or "").strip()
                gcal_declared = str(ws.cell(row=r, column=15).value or "").strip()
                is_confirmed = (
                    status_u == "ชำระแล้ว"
                    or bool(gcal_t)
                    or bool(gcal_declared)
                    or status_literal.lower() in ("confirmed", "ยืนยัน", "ชำระแล้ว")
                )
                return {
                    "found": True,
                    "sheet": "Bookings",
                    "status": "confirmed" if is_confirmed else "pending",
                    "raw_status": status_u or status_literal
                }

    # 2. ชีต "Cust" (ย้ายมาจาก PendingCust เมื่อแอดมินยืนยันแล้ว)
    if "Cust" in wb.sheetnames:
        ws = wb["Cust"]
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(row=r, column=18).value or "").strip() == queue_id:
                pay_status = str(ws.cell(row=r, column=19).value or "").strip()
                status = str(ws.cell(row=r, column=5).value or "").strip()
                is_confirmed = (pay_status == "ชำระแล้ว") or (status == "ยืนยัน")
                return {
                    "found": True,
                    "sheet": "Cust",
                    "status": "confirmed" if is_confirmed else "pending",
                    "raw_status": pay_status or status
                }

    # 3. ชีต "PendingCust" (คิวจากแชท ที่ยังรอตรวจ)
    if "PendingCust" in wb.sheetnames:
        ws = wb["PendingCust"]
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(row=r, column=18).value or "").strip() == queue_id:
                pay_status = str(ws.cell(row=r, column=19).value or "").strip()
                status = str(ws.cell(row=r, column=5).value or "").strip()
                is_confirmed = (pay_status == "ชำระแล้ว") or (status == "ยืนยัน")
                return {
                    "found": True,
                    "sheet": "PendingCust",
                    "status": "confirmed" if is_confirmed else "pending",
                    "raw_status": pay_status or status
                }

    return {"found": False}


@app.route("/api/booking-status", methods=["GET"])
def booking_status():
    """
    ให้หน้าเว็บดึงสถานะการจองจริงมาแสดงตอนเปิดแท็บ "คิวของฉัน"
    แทนการเดาสถานะเองแบบ setTimeout ฝั่ง browser
    """
    queue_id = (request.args.get("queueId") or "").strip()
    if not queue_id:
        return jsonify({"found": False, "error": "missing queueId"}), 400
    result = _lookup_booking_status(queue_id)
    if result is None:
        return jsonify({"found": False, "error": "excel_unavailable"}), 503
    return jsonify(result)


EVENTS_LOG_PATH = os.path.join(LOG_DIR, "events_log.jsonl")
_events_log_lock = threading.Lock()

@app.route("/api/track-event", methods=["POST", "OPTIONS"])
def track_event():
    """
    รับ event สั้นๆ จากหน้าเว็บ (เช่น calculator_used, booking_tab_opened) เพื่อ
    เก็บสถิติ conversion funnel: มีคนใช้เครื่องคำนวณ/เปิดหน้าจองกี่คน เทียบกับ
    จำนวนแถวจองจริงในชีต Bookings/Cust ที่มีอยู่แล้ว จะได้ conversion rate คร่าวๆ

    หมายเหตุ: จงใจ "ไม่" เขียนลง MuHub_Customer_DB.xlsx เหมือนข้อมูลจองจริง เพราะ
    event นี้อาจถูกยิงถี่ (ทุกครั้งที่กดผูกดวง/เปิดแท็บจอง) การเปิด-ปิด workbook
    ลูกค้าจริงบ่อยๆ จะเพิ่มความเสี่ยง lock ไฟล์ชนกับตอนแอดมินเปิดไฟล์ดูอยู่ จึงเขียน
    ลงไฟล์ .jsonl แยกต่างหาก (เบากว่ามาก) แล้วให้แอดมินนำไปสรุปเป็นระยะแทน
    """
    if request.method == "OPTIONS":
        return "", 200

    guard_result = require_app_token_and_rate_limit("track_event")
    if guard_result is not None:
        return guard_result

    try:
        data = request.get_json(force=True) or {}
        event_name = str(data.get("event", ""))[:100].strip()
        session_id = str(data.get("sessionId", ""))[:100].strip()
        detail = data.get("detail", {})

        if not event_name:
            return jsonify({"success": False, "error": "missing_event"}), 400

        record = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "event": event_name,
            "sessionId": session_id,
            "detail": detail,
        }
        with _events_log_lock:
            with open(EVENTS_LOG_PATH, "a", encoding="utf-8") as f:
                f.write(json.dumps(record, ensure_ascii=False) + "\n")

        return jsonify({"success": True})
    except Exception as e:
        print(f"[TrackEvent] Error (non-critical): {e}")
        return jsonify({"success": False, "error": str(e)})

@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({
        "status": "ok", 
        "excel": os.path.exists(EXCEL_PATH),
        "excel_path": EXCEL_PATH,
        "line_token_set": bool(LINE_CHANNEL_ACCESS_TOKEN),
        "line_secret_set": bool(LINE_CHANNEL_SECRET)
    })

if __name__ == "__main__":
    print("=" * 55)
    print("  MuHub Excel & LINE Server  —  http://localhost:5001")
    print(f"  Excel Path: {EXCEL_PATH}")
    print(f"  Excel Exists: {os.path.exists(EXCEL_PATH)}")
    print("=" * 55)
    
    # ดึงคิวค้างมาเคลียร์ก่อนรันระบบ
    if os.path.exists(EXCEL_PATH):
        try:
            wb = load_workbook(EXCEL_PATH)
            if "Cust" in wb.sheetnames:
                ws = wb["Cust"]
                process_pending_queue(wb, ws)
                wb.save(EXCEL_PATH)
        except Exception as e:
            print(f"[Startup] Warning: Failed to process queue on startup: {e}")
            
    # สตาร์ทเทรดเบื้องหลังสำหรับตรวจจับความเคลื่อนไหวและยืนยันคิวจาก Excel อัตโนมัติ
    import threading
    monitor_thread = threading.Thread(target=excel_monitor_thread_func, daemon=True)
    monitor_thread.start()
    print("[Startup] Excel modification monitor thread started.")

    app.run(port=PORT, debug=False)
