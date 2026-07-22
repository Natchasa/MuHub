"""
muhub_setup_lineid_sheet.py
รันครั้งเดียวเพื่อ rename sheet "LINE User IDs" → "Cust"
และ redesign ให้มี header styling เหมือน Customer DB
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = r"G:\My Drive\Customer\MuHub_Customer_DB.xlsx"
SHEET_NAME = "Cust"
OLD_SHEET_NAME = "LINE User IDs"

COLOR_HEADER_BG   = "3E2723"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_ROW_ODD     = "FFF8F5"
COLOR_ROW_EVEN    = "FFFFFF"
COLOR_BORDER      = "D7B899"
COLOR_TITLE_BG    = "6D4C41"

COLUMNS = [
    ("ลำดับ",        8),
    ("วันที่บันทึก", 18),
    ("Display Name", 22),
    ("LINE User ID", 28),
    ("สถานะ",        14),
    ("หมายเหตุ",     30),
]

def make_border(color=COLOR_BORDER):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)

def style_header_cell(cell, text):
    cell.value = text
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    cell.font = Font(name="Angsana New", bold=True, color=COLOR_HEADER_FONT, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border("5D4037")

def style_data_cell(cell, row_num):
    bg = COLOR_ROW_ODD if row_num % 2 == 1 else COLOR_ROW_EVEN
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Angsana New", size=13)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = make_border()

def setup_sheet():
    print("📂 เปิดไฟล์ Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # เก็บข้อมูลเดิมจากทั้ง "LINE User IDs" และ "Cust" (ถ้ามี)
    existing_data = []
    for sname in [OLD_SHEET_NAME, SHEET_NAME]:
        if sname in wb.sheetnames:
            old_ws = wb[sname]
            for row in old_ws.iter_rows(min_row=3, values_only=True):
                if any(v for v in row):
                    existing_data.append(row)
            del wb[sname]
            print(f"  ↳ ลบ sheet เดิม '{sname}' พบข้อมูล {len(existing_data)} แถว")

    ws = wb.create_sheet(SHEET_NAME)

    # Title row
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:F1")
    tc = ws["A1"]
    tc.value = "LINE User IDs — MuHub Customer"
    tc.fill = PatternFill("solid", fgColor=COLOR_TITLE_BG)
    tc.font = Font(name="Angsana New", bold=True, color="FFFFFF", size=16)
    tc.alignment = Alignment(horizontal="center", vertical="center")

    # Header row
    ws.row_dimensions[2].height = 28
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx)
        style_header_cell(cell, col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Migrate ข้อมูล
    for i, old_row in enumerate(existing_data):
        r = 3 + i
        ws.row_dimensions[r].height = 22
        # รองรับทั้ง format เก่า (3 cols) และ format ใหม่ (6 cols)
        if len(old_row) >= 6:
            values = list(old_row[:6])
            values[0] = i + 1  # เรียงลำดับใหม่
        else:
            values = [i + 1, old_row[0] or "", old_row[1] or "", old_row[2] or "", "ใหม่", ""]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            style_data_cell(cell, i)
        print(f"  ✅ Migrated: {values[2]} | {values[3]}")

    ws.freeze_panes = "A3"
    wb.save(EXCEL_PATH)
    print(f"\n✅ Sheet '{SHEET_NAME}' พร้อมใช้งานแล้วค่ะ!")
    print(f"📁 บันทึกที่: {EXCEL_PATH}")

if __name__ == "__main__":
    setup_sheet()
