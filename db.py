import sqlite3
import os
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "muhub.db")

# Excel headers definitions
BOOKINGS_HEADERS = [
    "Timestamp", "Queue ID", "Customer Name", "LINE ID",
    "Astrologer", "Service", "Date", "Time Slot", "Questions",
    "Birth Date", "Birth Hour", "Birth Min",
    "Birth Country", "Birth City", "GCal Event ID", "Status"
]

CUST_HEADERS = [
    "Customer ID", "วันที่บันทึก", "Display Name", "LINE User ID", "สถานะ", 
    "วันที่จอง", "ชื่อลูกค้า", "Line ID", "วันเกิด (พ.ศ.)", "เวลาเกิด", 
    "ประเทศ", "จังหวัด", "แพคเกจ/บริการ", "หมอดู", "วันนัดหมาย", 
    "ช่วงเวลานัด", "คำถามพิเศษ", "เลขอ้างอิง", "สถานะการชำระ", "ประเภทลูกค้า", 
    "จำนวนครั้งที่จอง", "หมายเหตุ"
]

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    """Initialize SQLite tables if they do not exist."""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 1. Bookings table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS bookings (
        timestamp TEXT,
        queue_id TEXT PRIMARY KEY,
        name TEXT,
        line_id TEXT,
        astrologer TEXT,
        service TEXT,
        date TEXT,
        slot TEXT,
        questions TEXT,
        birth_date TEXT,
        birth_hour TEXT,
        birth_min TEXT,
        birth_country TEXT,
        birth_city TEXT,
        gcal_event_id TEXT,
        status TEXT
    )
    """)
    
    # 2. Customers table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS customers (
        seq INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp TEXT,
        display_name TEXT,
        user_id TEXT UNIQUE,
        status TEXT,
        booking_date TEXT,
        customer_name TEXT,
        line_id TEXT,
        birth_date TEXT,
        birth_time TEXT,
        birth_country TEXT,
        birth_city TEXT,
        package TEXT,
        astrologer TEXT,
        booking_date_appointment TEXT,
        booking_slot TEXT,
        questions TEXT,
        queue_id TEXT,
        payment_status TEXT,
        customer_type TEXT,
        booking_count INTEGER,
        notes TEXT
    )
    """)
    
    # 3. Pending Customers table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS pending_customers (
        seq INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp TEXT,
        display_name TEXT,
        user_id TEXT UNIQUE,
        status TEXT,
        booking_date TEXT,
        customer_name TEXT,
        line_id TEXT,
        birth_date TEXT,
        birth_time TEXT,
        birth_country TEXT,
        birth_city TEXT,
        package TEXT,
        astrologer TEXT,
        booking_date_appointment TEXT,
        booking_slot TEXT,
        questions TEXT,
        queue_id TEXT,
        payment_status TEXT,
        customer_type TEXT,
        booking_count INTEGER,
        notes TEXT
    )
    """)
    
    # 4. Promo code used table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS promo_used (
        code TEXT,
        line_id TEXT,
        PRIMARY KEY (code, line_id)
    )
    """)
    
    # 5. Events log table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS events_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp TEXT,
        event_name TEXT,
        session_id TEXT,
        detail TEXT
    )
    """)
    
    conn.commit()
    conn.close()
    print("[DB] Tables initialized successfully.")

def import_excel_to_sqlite(excel_path):
    """Import existing Excel data to SQLite database (run once on startup if db is new)."""
    if not os.path.exists(excel_path):
        print(f"[DB Migration] Excel file not found at {excel_path}. Skipping initial import.")
        return False
        
    print(f"[DB Migration] Migrating data from {excel_path} to SQLite...")
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Helper function to clear table and import
        def migrate_sheet(sheet_name, table_name, start_row, headers_list, insert_query, row_mapper_func):
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                cursor.execute(f"DELETE FROM {table_name}")
                count = 0
                for r in range(start_row, ws.max_row + 1):
                    row_vals = [ws.cell(row=r, column=c).value for c in range(1, len(headers_list) + 1)]
                    # Check if row is empty (especially columns like Queue ID / Name / LINE User ID)
                    if not any(row_vals):
                        continue
                    mapped_vals = row_mapper_func(row_vals)
                    try:
                        cursor.execute(insert_query, mapped_vals)
                        count += 1
                    except sqlite3.IntegrityError as ie:
                        # Handle duplicate PRIMARY KEY / UNIQUE constraints gracefully
                        # e.g., if there's multiple entries for a user, update it
                        if "user_id" in insert_query or "LINE User ID" in sheet_name or "Cust" in sheet_name:
                            # Update existing row instead of throwing error
                            u_id = mapped_vals[3] # User ID is at index 3 in mapped values
                            if u_id:
                                # Construct an update query based on table columns
                                cols = ["timestamp", "display_name", "status", "booking_date", "customer_name", "line_id",
                                        "birth_date", "birth_time", "birth_country", "birth_city", "package", "astrologer",
                                        "booking_date_appointment", "booking_slot", "questions", "queue_id", "payment_status",
                                        "customer_type", "booking_count", "notes"]
                                set_clause = ", ".join([f"{col} = ?" for col in cols])
                                # Remove seq from updates, user_id is the condition
                                vals_to_update = list(mapped_vals[1:3]) + list(mapped_vals[4:]) + [u_id]
                                cursor.execute(f"UPDATE {table_name} SET {set_clause} WHERE user_id = ?", vals_to_update)
                                count += 1
                        else:
                            print(f"[DB Migration] Skip duplicates in {sheet_name}: {ie}")
                print(f"[DB Migration] Migrated {count} rows to {table_name}")
        
        # 1. Bookings mapping
        def map_booking(r):
            # Excel columns: 1:Timestamp, 2:QueueID, 3:Name, 4:LineID, 5:Astrologer, 6:Service, 7:Date, 8:Slot, 9:Questions, 
            # 10:BirthDate, 11:BirthHour, 12:BirthMin, 13:BirthCountry, 14:BirthCity, 15:GCalID, 16:Status
            # Wait, our Excel file might have columns at different offsets.
            # Let's map safely based on header index or fallback
            return (
                str(r[0]) if r[0] else "",
                str(r[1]) if r[1] else "",
                str(r[2]) if r[2] else "",
                str(r[3]) if r[3] else "",
                str(r[4]) if r[4] else "",
                str(r[5]) if r[5] else "",
                str(r[6]) if r[6] else "",
                str(r[7]) if r[7] else "",
                str(r[8]) if r[8] else "",
                str(r[9]) if r[9] else "",
                str(r[10]) if r[10] else "",
                str(r[11]) if r[11] else "",
                str(r[12]) if r[12] else "",
                str(r[13]) if r[13] else "",
                str(r[14]) if r[14] else "",
                str(r[15]) if r[15] else "pending"
            )
            
        bookings_insert = """
        INSERT INTO bookings (
            timestamp, queue_id, name, line_id, astrologer, service, date, slot, questions,
            birth_date, birth_hour, birth_min, birth_country, birth_city, gcal_event_id, status
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        migrate_sheet("Bookings", "bookings", 2, BOOKINGS_HEADERS, bookings_insert, map_booking)
        
        # 2. Cust and PendingCust mapping
        def map_cust(r):
            # Excel columns: 1:ID, 2:Timestamp, 3:DisplayName, 4:LINEUserID, 5:Status, 6:BookingDate, 7:CustName, 8:LineID,
            # 9:BirthDate, 10:BirthTime, 11:Country, 12:City, 13:Service, 14:Astrologer, 15:DateAppt, 16:Slot, 17:Questions,
            # 18:QueueID, 19:PaymentStatus, 20:Type, 21:Count, 22:Notes
            # We skip seq (column 1) because SQLite auto-increments it
            # We ensure values are parsed
            return (
                None, # SQLite seq
                str(r[1]) if r[1] else "",
                str(r[2]) if r[2] else "",
                str(r[3]) if r[3] else "",
                str(r[4]) if r[4] else "ใหม่",
                str(r[5]) if r[5] else "",
                str(r[6]) if r[6] else "",
                str(r[7]) if r[7] else "",
                str(r[8]) if r[8] else "",
                str(r[9]) if r[9] else "",
                str(r[10]) if r[10] else "",
                str(r[11]) if r[11] else "",
                str(r[12]) if r[12] else "",
                str(r[13]) if r[13] else "",
                str(r[14]) if r[14] else "",
                str(r[15]) if r[15] else "",
                str(r[16]) if r[16] else "",
                str(r[17]) if r[17] else "",
                str(r[18]) if r[18] else "",
                str(r[19]) if r[19] else "",
                int(r[20]) if r[20] and str(r[20]).isdigit() else 0,
                str(r[21]) if r[21] else ""
            )
            
        cust_insert = """
        INSERT INTO customers (
            seq, timestamp, display_name, user_id, status, booking_date, customer_name, line_id,
            birth_date, birth_time, birth_country, birth_city, package, astrologer,
            booking_date_appointment, booking_slot, questions, queue_id, payment_status,
            customer_type, booking_count, notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        migrate_sheet("Cust", "customers", 3, CUST_HEADERS, cust_insert, map_cust)
        migrate_sheet("PendingCust", "pending_customers", 3, CUST_HEADERS, cust_insert, map_cust)
        
        conn.commit()
        conn.close()
        wb.close()
        print("[DB Migration] Completed Excel migration to SQLite!")
        return True
    except Exception as e:
        print(f"[DB Migration] Fatal error migrating Excel to SQLite: {e}")
        return False

def export_sqlite_to_excel(excel_path):
    """Query SQLite tables and build a formatted Excel spreadsheet."""
    wb = openpyxl.Workbook()
    
    # Border & Fills styling
    thin_border = Border(
        left=Side(style='thin', color='D4AF37'),
        right=Side(style='thin', color='D4AF37'),
        top=Side(style='thin', color='D4AF37'),
        bottom=Side(style='thin', color='D4AF37')
    )
    header_fill = PatternFill(start_color="3E2723", end_color="3E2723", fill_type="solid")
    header_font = Font(name="Sarabun", size=11, bold=True, color="FFFFFF")
    
    # 1. Export Bookings
    ws_bookings = wb.active
    ws_bookings.title = "Bookings"
    ws_bookings.append(BOOKINGS_HEADERS)
    # Style headers
    for col in range(1, len(BOOKINGS_HEADERS) + 1):
        cell = ws_bookings.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM bookings ORDER BY timestamp DESC")
    bookings_rows = cursor.fetchall()
    
    for brow in bookings_rows:
        row_vals = [brow[col] for col in brow.keys()]
        ws_bookings.append(row_vals)
        
    # Auto fit column widths
    for col in ws_bookings.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws_bookings.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)
        
    # 2. Export Cust & PendingCust Sheets
    def export_cust_sheet(table_name, sheet_name):
        ws = wb.create_sheet(sheet_name)
        
        # Merged title row
        ws.row_dimensions[1].height = 30
        ws.merge_cells(f"A1:{get_column_letter(len(CUST_HEADERS))}1")
        tc = ws["A1"]
        tc.value = f"LINE Users Database — MuHub {sheet_name}"
        tc.fill = PatternFill("solid", fgColor="6D4C41")
        tc.font = Font(name="Sarabun", bold=True, color="FFFFFF", size=14)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.row_dimensions[2].height = 28
        for col_idx, h_name in enumerate(CUST_HEADERS, start=1):
            cell = ws.cell(row=2, column=col_idx, value=h_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        cursor.execute(f"SELECT * FROM {table_name} ORDER BY seq ASC")
        cust_rows = cursor.fetchall()
        
        for idx, crow in enumerate(cust_rows, start=3):
            # Column 1 is seq formula (Customer ID)
            ws.cell(row=idx, column=1).value = f'=IF(C{idx}<>"",TEXT(ROW()-3,"0000"),"")'
            ws.cell(row=idx, column=2).value = crow["timestamp"]
            ws.cell(row=idx, column=3).value = crow["display_name"]
            ws.cell(row=idx, column=4).value = crow["user_id"]
            ws.cell(row=idx, column=5).value = crow["status"]
            ws.cell(row=idx, column=6).value = crow["booking_date"]
            ws.cell(row=idx, column=7).value = crow["customer_name"]
            ws.cell(row=idx, column=8).value = crow["line_id"]
            ws.cell(row=idx, column=9).value = crow["birth_date"]
            ws.cell(row=idx, column=10).value = crow["birth_time"]
            ws.cell(row=idx, column=11).value = crow["birth_country"]
            ws.cell(row=idx, column=12).value = crow["birth_city"]
            ws.cell(row=idx, column=13).value = crow["package"]
            ws.cell(row=idx, column=14).value = crow["astrologer"]
            ws.cell(row=idx, column=15).value = crow["booking_date_appointment"]
            ws.cell(row=idx, column=16).value = crow["booking_slot"]
            ws.cell(row=idx, column=17).value = crow["questions"]
            ws.cell(row=idx, column=18).value = crow["queue_id"]
            ws.cell(row=idx, column=19).value = crow["payment_status"]
            
            # Formulas
            ws.cell(row=idx, column=20).value = f'=IF(U{idx}>1,"ลูกค้าเก่า","ลูกค้าใหม่")'
            ws.cell(row=idx, column=21).value = f'=COUNTIF(D:D, D{idx})'
            ws.cell(row=idx, column=22).value = crow["notes"]
            
            # Style data cells
            for col_idx in range(1, len(CUST_HEADERS) + 1):
                cell = ws.cell(row=idx, column=col_idx)
                cell.font = Font(name="Sarabun", size=10)
                cell.border = thin_border
                
        # Auto-fit columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)
            
    export_cust_sheet("customers", "Cust")
    export_cust_sheet("pending_customers", "PendingCust")
    
    conn.close()
    
    # Save safely
    wb.save(excel_path)
    wb.close()
    print(f"[DB] Exported SQLite database to Excel: {excel_path}")

# Initialize database on load
init_db()
