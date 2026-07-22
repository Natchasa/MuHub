import os
import sys
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
import csv
import math
import datetime
import zipfile
import json
import logging
from logging.handlers import RotatingFileHandler
import urllib.request
import urllib.parse
import hmac
import hashlib
import base64
import time
import threading
from collections import defaultdict, deque
from typing import Dict, Any, List, Optional
import requests
from fastapi import FastAPI, HTTPException, Query, Request, UploadFile, File
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from timezonefinder import TimezoneFinder
import pytz
from pydantic import BaseModel
import openpyxl

import db

# Setup Directories and Logging
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_DIR = os.path.join(BASE_DIR, "logs")
os.makedirs(LOG_DIR, exist_ok=True)

_server_logger = logging.getLogger("muhub_server")
_server_logger.setLevel(logging.INFO)
_server_logger.propagate = False
if not _server_logger.handlers:
    _log_handler = RotatingFileHandler(
        os.path.join(LOG_DIR, "server.log"), maxBytes=5 * 1024 * 1024, backupCount=5, encoding="utf-8"
    )
    _log_handler.setFormatter(logging.Formatter("%(asctime)s %(message)s"))
    _server_logger.addHandler(_log_handler)

class _StreamToLogger:
    def __init__(self, logger, level, original_stream):
        self.logger = logger
        self.level = level
        self.original_stream = original_stream
        self._buffer = ""

    def write(self, message):
        self.original_stream.write(message)
        self._buffer += message
        while "\n" in self._buffer:
            line, self._buffer = self._buffer.split("\n", 1)
            if line.strip():
                self.logger.log(self.level, line)

    def flush(self):
        self.original_stream.flush()

    def isatty(self):
        return hasattr(self.original_stream, 'isatty') and self.original_stream.isatty()

    def fileno(self):
        if hasattr(self.original_stream, 'fileno'):
            return self.original_stream.fileno()
        raise OSError("Stream has no fileno")

sys.stdout = _StreamToLogger(_server_logger, logging.INFO, sys.stdout)
sys.stderr = _StreamToLogger(_server_logger, logging.ERROR, sys.stderr)

# Load configuration (Fallback to Env Variables)
CONFIG = {}
config_path = os.path.join(BASE_DIR, "config.json")
if os.path.exists(config_path):
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            CONFIG = json.load(f)
    except Exception as e:
        print(f"Error reading config.json: {e}")

LINE_CHANNEL_ACCESS_TOKEN = os.environ.get("LINE_CHANNEL_ACCESS_TOKEN") or CONFIG.get("LINE_CHANNEL_ACCESS_TOKEN", "")
LINE_CHANNEL_SECRET = os.environ.get("LINE_CHANNEL_SECRET") or CONFIG.get("LINE_CHANNEL_SECRET", "")
LINE_ADMIN_USER_ID = os.environ.get("LINE_ADMIN_USER_ID") or CONFIG.get("LINE_ADMIN_USER_ID", "")
EXCEL_PATH = os.environ.get("EXCEL_PATH") or CONFIG.get("EXCEL_PATH", r"G:\My Drive\Customer\MuHub_Customer_DB.xlsx")
GOOGLE_SCRIPT_URL = os.environ.get("GOOGLE_SCRIPT_URL") or CONFIG.get("GOOGLE_SCRIPT_URL", "")
SLIPOK_API_KEY = os.environ.get("SLIPOK_API_KEY") or CONFIG.get("SLIPOK_API_KEY", "")
SLIPOK_BRANCH_ID = os.environ.get("SLIPOK_BRANCH_ID") or CONFIG.get("SLIPOK_BRANCH_ID", "")
APP_API_TOKEN = os.environ.get("APP_API_TOKEN") or CONFIG.get("APP_API_TOKEN", "")
PORT = int(os.environ.get("PORT") or CONFIG.get("PORT", 3737))

print(f"[Config] Consolidated server configuration loaded.")
print(f"[Config] Local Excel Path: {EXCEL_PATH}")
print(f"[Config] LINE Token configured: {bool(LINE_CHANNEL_ACCESS_TOKEN)}")
print(f"[Config] SlipOK configured: {bool(SLIPOK_API_KEY)}")
print(f"[Config] App API Token configured: {bool(APP_API_TOKEN)}")

# Run SQLite Database migration on startup if Excel file exists
db.import_excel_to_sqlite(EXCEL_PATH)

# Initialize Timezone Finder
tf = TimezoneFinder()

# Initialize FastAPI
app = FastAPI(
    title="MuHub Consolidated Server",
    description="Unified API server serving MuHub backend, booking, and LINE webhook systems."
)

# Enable CORS (Configurable via ALLOWED_ORIGINS env/config)
allowed_origins_config = CONFIG.get("ALLOWED_ORIGINS", ["*"])
if isinstance(allowed_origins_config, str):
    allowed_origins_config = [o.strip() for o in allowed_origins_config.split(",") if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins_config,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Block sensitive files middleware
BLOCKED_PATH_SEGMENTS = [
    "config.json",
    "non-github",
    ".env",
    ".git",
    "__pycache__",
    ".venv",
    ".claude",
    "muhub.db"
]

@app.middleware("http")
async def block_sensitive_files(request: Request, call_next):
    path = request.url.path.lower()
    if any(seg in path for seg in BLOCKED_PATH_SEGMENTS):
        return JSONResponse(status_code=404, content={"detail": "Not Found"})
    return await call_next(request)

@app.middleware("http")
async def catch_unhandled_errors(request: Request, call_next):
    try:
        return await call_next(request)
    except Exception as e:
        _server_logger.exception(f"[UnhandledError] {request.method} {request.url.path}: {e}")
        return JSONResponse(
            status_code=500,
            content={"detail": "เกิดข้อผิดพลาดที่ไม่คาดคิดในระบบ กรุณาลองใหม่อีกครั้ง"}
        )

@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
    return response

# Rate Limiting Utilities (consolidated from Flask)
_rate_limit_buckets = defaultdict(deque)
_rate_limit_lock = threading.Lock()
RATE_LIMIT_MAX_REQUESTS = 12
RATE_LIMIT_WINDOW_SECONDS = 60

def _get_client_ip(request: Request) -> str:
    forwarded = request.headers.get("X-Forwarded-For", "")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"

def check_rate_limit(bucket_key: str) -> bool:
    now = time.time()
    with _rate_limit_lock:
        q = _rate_limit_buckets[bucket_key]
        while q and now - q[0] > RATE_LIMIT_WINDOW_SECONDS:
            q.popleft()
        if len(q) >= RATE_LIMIT_MAX_REQUESTS:
            return False
        q.append(now)
        return True

def verify_app_token_and_rate_limit(request: Request, endpoint_name: str):
    client_ip = _get_client_ip(request)
    if not check_rate_limit(f"{endpoint_name}:{client_ip}"):
        print(f"[Security] Rate limit exceeded for {endpoint_name} from {client_ip}")
        raise HTTPException(status_code=429, detail="ส่งคำขอถี่เกินไป กรุณาลองใหม่ภายหลัง")

    if APP_API_TOKEN:
        provided = request.headers.get("X-App-Token", "")
        if provided != APP_API_TOKEN:
            print(f"[Security] Invalid/missing app token for {endpoint_name} from {client_ip}")
            raise HTTPException(status_code=401, detail="Unauthorized")

# Prevent Formula Injection
_DANGEROUS_LEADING_CHARS = ("=", "+", "-", "@", "\t", "\r")
def sanitize_excel_value(val):
    if isinstance(val, str) and val.startswith(_DANGEROUS_LEADING_CHARS):
        return "'" + val
    return val

# ── ASTRONOMY / EPHEMERIS SYSTEM ───────────────────────────────────────
CSV_PATH = os.path.join(BASE_DIR, "thai-astrology-ephemeris", "thai_ephemeris_1900_2200.csv")
EPHEMERIS_DATA: Dict[str, Dict[str, Any]] = {}

def load_ephemeris_csv():
    global CSV_PATH
    if not os.path.exists(CSV_PATH):
        zip_path = CSV_PATH.replace(".csv", ".zip")
        if os.path.exists(zip_path):
            print(f"CSV database not found. Extracting from {zip_path}...")
            try:
                extract_dir = os.path.dirname(CSV_PATH)
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    zip_ref.extractall(extract_dir)
                print("Successfully extracted CSV database!")
            except Exception as e:
                print(f"Error extracting ZIP file: {e}")
                return
        else:
            print(f"Warning: Ephemeris CSV not found at {CSV_PATH}. Fallback mode only.")
            return
        
    print(f"Loading ephemeris CSV from {CSV_PATH}...")
    start_time = datetime.datetime.now()
    with open(CSV_PATH, mode="r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            EPHEMERIS_DATA[row["date"]] = row
    elapsed = (datetime.datetime.now() - start_time).total_seconds()
    print(f"Successfully loaded {len(EPHEMERIS_DATA)} rows in {elapsed:.2f} seconds.")

load_ephemeris_csv()

def mod360(x: float) -> float:
    return ((x % 360) + 360) % 360

def to_rad(d: float) -> float:
    return d * math.pi / 180

def to_deg(r: float) -> float:
    return r * 180 / math.pi

def gmst(jd: float) -> float:
    T = (jd - 2451545.0) / 36525.0
    return mod360(280.46061837 + 360.98564736629 * (jd - 2451545.0) + 0.000387933 * T * T - T * T * T / 38710000.0)

def obliq(jd: float) -> float:
    return 23.439291111 - 0.013004167 * (jd - 2451545.0) / 36525.0

def calculate_lagna(jd: float, lat: float, lon: float, ayanamsa: float) -> float:
    ramc = mod360(gmst(jd) + lon)
    e = to_rad(obliq(jd))
    r = to_rad(ramc)
    l = to_rad(lat)
    num = -math.cos(r)
    den = math.sin(r) * math.cos(e) + math.tan(l) * math.sin(e)
    asc = to_deg(math.atan2(num, den))
    return mod360(asc + 180 - ayanamsa)

@app.get("/api/calculate")
def calculate_from_csv(
    date: str = Query(..., description="Date in YYYY-MM-DD format"),
    time: str = Query("07:00", description="Local time in HH:MM format"),
    lat: float = Query(13.7563, description="Latitude"),
    lon: float = Query(100.5018, description="Longitude"),
    tz: float = Query(None, description="Optional timezone offset in hours")
):
    if not EPHEMERIS_DATA:
        raise HTTPException(status_code=503, detail="Ephemeris CSV database is not loaded.")
        
    try:
        dt_local = datetime.datetime.strptime(f"{date} {time}", "%Y-%m-%d %H:%M")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date or time format. Use YYYY-MM-DD and HH:MM.")
        
    if tz is None:
        try:
            tz_name = tf.timezone_at(lng=lon, lat=lat)
            if tz_name:
                timezone_obj = pytz.timezone(tz_name)
                try:
                    loc_dt = timezone_obj.localize(dt_local, is_dst=None)
                except Exception:
                    loc_dt = timezone_obj.localize(dt_local)
                tz = loc_dt.utcoffset().total_seconds() / 3600.0
            else:
                tz = round(lon / 15.0)
        except Exception as e:
            print(f"Error calculating timezone offset: {e}")
            tz = round(lon / 15.0)
            
    dt_utc = dt_local - datetime.timedelta(hours=tz)
    row_current = EPHEMERIS_DATA.get(date)
    if not row_current:
        raise HTTPException(status_code=400, detail=f"Date {date} is out of the database range (1900-2200).")
        
    try:
        sunrise_curr = datetime.datetime.fromisoformat(row_current["sunrise_utc"]).replace(tzinfo=None)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error parsing database row: {e}")
        
    if dt_utc >= sunrise_curr:
        date_next = (dt_local.date() + datetime.timedelta(days=1)).strftime("%Y-%m-%d")
        row_start = row_current
        row_end = EPHEMERIS_DATA.get(date_next)
    else:
        date_prev = (dt_local.date() - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
        row_start = EPHEMERIS_DATA.get(date_prev)
        row_end = row_current
        
    if not row_start or not row_end:
        raise HTTPException(status_code=400, detail="Bracketing dates are out of database range.")
        
    S_start = datetime.datetime.fromisoformat(row_start["sunrise_utc"]).replace(tzinfo=None)
    S_end = datetime.datetime.fromisoformat(row_end["sunrise_utc"]).replace(tzinfo=None)
    
    total_seconds = (S_end - S_start).total_seconds()
    elapsed_seconds = (dt_utc - S_start).total_seconds()
    
    fraction = 0.0 if total_seconds == 0 else elapsed_seconds / total_seconds
    
    jd_start = float(row_start["julian_day"])
    jd_end = float(row_end["julian_day"])
    julian_day = jd_start + fraction * (jd_end - jd_start)
    
    ay_start = float(row_start["ayanamsa"])
    ay_end = float(row_end["ayanamsa"])
    ay_diff = ay_end - ay_start
    if ay_diff < -180: ay_diff += 360
    elif ay_diff > 180: ay_diff -= 360
    ayanamsa = mod360(ay_start + fraction * ay_diff)
    
    planets = [
        "sun", "moon", "mars", "mercury", "jupiter", "venus", "saturn",
        "rahu", "ketu", "thai_ketu", "uranus", "neptune", "pluto"
    ]
    
    pos = {}
    retro = {}
    
    for p in planets:
        lon_start = float(row_start[f"{p}_longitude"])
        lon_end = float(row_end[f"{p}_longitude"])
        
        diff = lon_end - lon_start
        if diff < -180: diff += 360
        elif diff > 180: diff -= 360
            
        lon_interp = mod360(lon_start + fraction * diff)
        pos[p] = lon_interp
        
        if p == "lagna":
            retro[p] = ""
        else:
            if diff < 0:
                retro[p] = "พักร์"
            else:
                slow_thresholds = {
                    "sun": 0.96, "moon": 12.2, "mars": 0.15, "mercury": 0.40,
                    "jupiter": 0.04, "venus": 0.40, "saturn": 0.015, "rahu": 0.04,
                    "ketu": 0.04, "thai_ketu": 0.50, "uranus": 0.005, "neptune": 0.002,
                    "pluto": 0.0015
                }
                fast_thresholds = {
                    "sun": 1.01, "moon": 14.0, "mars": 0.60, "mercury": 1.30,
                    "jupiter": 0.10, "venus": 1.22, "saturn": 0.045, "rahu": 0.06,
                    "ketu": 0.06, "thai_ketu": 0.55, "uranus": 0.015, "neptune": 0.008,
                    "pluto": 0.005
                }
                
                if diff <= slow_thresholds.get(p, 0.01):
                    retro[p] = "มนต์"
                elif diff >= fast_thresholds.get(p, 1.5):
                    retro[p] = "เสริด"
                else:
                    retro[p] = ""
            
    lagna_lon = calculate_lagna(julian_day, lat, lon, ayanamsa)
    pos["lagna"] = lagna_lon
    retro["lagna"] = ""
    
    return {
        "pos": pos,
        "retro": retro,
        "ayanamsa": ayanamsa,
        "julian_day": julian_day,
        "timezone_offset": tz,
        "source": "database_csv"
    }

# ── CALENDAR ENDPOINTS ─────────────────────────────────────────────────
@app.get("/api/calendar/busy")
async def get_busy_slots(date: str):
    api_key = CONFIG.get("GOOGLE_CALENDAR_API_KEY", "")
    if not api_key:
        try:
            day = int(date.split("-")[2])
        except Exception:
            day = 1
        return {"busy_slots": [{"startHour": day % 8, "endHour": (day % 8) + 1}, {"startHour": (day + 3) % 8, "endHour": ((day + 3) % 8) + 1}], "simulated": True}
        
    calendar_id = "muhub54@gmail.com"
    time_min = f"{date}T00:00:00Z"
    time_max = f"{date}T23:59:59Z"
    url = f"https://www.googleapis.com/calendar/v3/calendars/{urllib.parse.quote(calendar_id)}/events?key={api_key}&timeMin={time_min}&timeMax={time_max}&singleEvents=true"
    
    try:
        req = urllib.request.Request(url, method="GET")
        with urllib.request.urlopen(req, timeout=10) as response:
            data = json.loads(response.read().decode("utf-8"))
            
        busy_slots = []
        for event in data.get("items", []):
            start_val = event.get("start", {}).get("dateTime") or event.get("start", {}).get("date")
            end_val = event.get("end", {}).get("dateTime") or event.get("end", {}).get("date")
            if not start_val or not end_val:
                continue
            try:
                if "T" not in start_val:
                    busy_slots.append({"startHour": 0.0, "endHour": 24.0})
                    continue
                
                dt_start = datetime.datetime.fromisoformat(start_val.replace("Z", "+00:00"))
                dt_end = datetime.datetime.fromisoformat(end_val.replace("Z", "+00:00"))
                
                tz_thai = datetime.timezone(datetime.timedelta(hours=7))
                if dt_start.tzinfo is not None:
                    dt_start = dt_start.astimezone(tz_thai)
                if dt_end.tzinfo is not None:
                    dt_end = dt_end.astimezone(tz_thai)
                    
                start_h = dt_start.hour + dt_start.minute / 60.0
                end_h = dt_end.hour + dt_end.minute / 60.0
                busy_slots.append({"startHour": start_h, "endHour": end_h})
            except Exception as ex:
                print(f"Error parsing event: {ex} (start={start_val}, end={end_val})")
        return {"busy_slots": busy_slots, "simulated": False}
    except Exception as e:
        try:
            day = int(date.split("-")[2])
        except Exception:
            day = 1
        return {"busy_slots": [{"startHour": day % 8, "endHour": (day % 8) + 1}, {"startHour": (day + 3) % 8, "endHour": ((day + 3) % 8) + 1}], "simulated": True, "error": str(e)}

class BlockPayload(BaseModel):
    dateStr: str
    slotStr: str
    name: str
    lineId: str
    serviceName: str
    astrologerName: str
    questions: str = ""

@app.post("/api/calendar/block")
async def block_calendar_event(payload: BlockPayload, request: Request):
    client_ip = _get_client_ip(request)
    if not check_rate_limit(f"block_calendar:{client_ip}"):
        raise HTTPException(status_code=429, detail="ส่งคำขอจองคิวถี่เกินไป กรุณาลองใหม่ภายหลัง")

    safe_name = payload.name[:100].strip()
    safe_line_id = payload.lineId[:50].strip()
    safe_service = payload.serviceName[:100].strip()
    safe_astrologer = payload.astrologerName[:100].strip()
    safe_questions = payload.questions[:500].strip()

    if not GOOGLE_SCRIPT_URL:
        return {"success": True, "simulated": True, "data": {"eventId": f"mock-gcal-{datetime.datetime.now().microsecond}"}}
        
    parts = payload.slotStr.split("-")
    start_part = parts[0].strip()
    end_part = parts[1].strip() if len(parts) > 1 else ""
    start_hour = start_part[:2]
    end_hour = end_part[:2] if end_part else f"{int(start_hour) + 1:02d}"
    end_minute = end_part[3:5] if end_part else "00"
    
    start_str = f"{payload.dateStr}T{start_hour}:00:00+07:00"
    end_str = f"{payload.dateStr}T{end_hour}:{end_minute}:00+07:00"
    
    gas_payload = {
        "action": "create",
        "summary": f"MuHub จองคิว: คุณ {safe_name} (Line: {safe_line_id or 'ไม่ระบุ'})",
        "description": f"บริการ: {safe_service}\nนักพยากรณ์: {safe_astrologer}\nLine ID: {safe_line_id or 'ไม่ระบุ'}\nคำถามที่สนใจ: {safe_questions or 'ไม่มี'}\nบันทึกจากระบบ MuHub Booking",
        "startTime": start_str,
        "endTime": end_str
    }
    
    try:
        data_json = json.dumps(gas_payload).encode("utf-8")
        req = urllib.request.Request(
            GOOGLE_SCRIPT_URL,
            data=data_json,
            headers={"Content-Type": "text/plain;charset=utf-8"},
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=15) as response:
            res_data = json.loads(response.read().decode("utf-8"))
        if res_data and res_data.get("success") is False:
            raise Exception(res_data.get("error") or "Google Apps Script error")
        return {"success": True, "simulated": False, "data": res_data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class DeletePayload(BaseModel):
    eventId: str

@app.post("/api/calendar/delete")
async def delete_calendar_event(payload: DeletePayload):
    if not GOOGLE_SCRIPT_URL or payload.eventId.startswith("mock-") or payload.eventId.startswith("mock_"):
        return {"success": True, "simulated": True, "message": "ลบกิจกรรมจำลองสำเร็จ"}
        
    gas_payload = {
        "action": "delete",
        "eventId": payload.eventId
    }
    
    try:
        data_json = json.dumps(gas_payload).encode("utf-8")
        req = urllib.request.Request(
            GOOGLE_SCRIPT_URL,
            data=data_json,
            headers={"Content-Type": "text/plain;charset=utf-8"},
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=15) as response:
            res_data = json.loads(response.read().decode("utf-8"))
        return {"success": True, "data": res_data}
    except Exception as e:
        return {"success": False, "error": str(e)}

ALLOWED_SLIP_MIMES = {"image/jpeg", "image/png", "image/webp", "image/jpg"}
MAX_BASE64_LENGTH = 7 * 1024 * 1024  # Max ~5MB raw file

class UploadPayload(BaseModel):
    base64Data: str
    mimeType: str
    filename: str

@app.post("/api/calendar/upload-slip")
async def upload_slip_event(payload: UploadPayload, request: Request):
    client_ip = _get_client_ip(request)
    if not check_rate_limit(f"upload_slip:{client_ip}"):
        raise HTTPException(status_code=429, detail="ส่งคำขออัปโหลดถี่เกินไป กรุณาลองใหม่ภายหลัง")

    clean_mime = payload.mimeType.strip().lower()
    if clean_mime not in ALLOWED_SLIP_MIMES:
        raise HTTPException(status_code=400, detail="ชนิดไฟล์ไม่ถูกต้อง รองรับเฉพาะไฟล์รูปภาพ (JPG, PNG, WEBP) เท่านั้น")

    if len(payload.base64Data) > MAX_BASE64_LENGTH:
        raise HTTPException(status_code=400, detail="ขนาดไฟล์ใหญ่เกินกำหนด (สูงสุด 5MB)")

    if not GOOGLE_SCRIPT_URL:
        return {"success": True, "simulated": True, "fileUrl": "https://drive.google.com/open?id=mock-drive-file-id"}
        
    gas_payload = {
        "action": "uploadSlip",
        "folderId": "1RtwEoH4ES9QJT2oB8JobUQw1rncnETYF",
        "filename": payload.filename[:100],
        "mimeType": clean_mime,
        "base64Data": payload.base64Data
    }
    
    try:
        data_json = json.dumps(gas_payload).encode("utf-8")
        req = urllib.request.Request(
            GOOGLE_SCRIPT_URL,
            data=data_json,
            headers={"Content-Type": "text/plain;charset=utf-8"},
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=30) as response:
            res_data = json.loads(response.read().decode("utf-8"))
        if res_data and res_data.get("success") is False:
            raise Exception(res_data.get("error") or "Google Apps Script error")
        return {"success": True, "simulated": False, "fileUrl": res_data.get("fileUrl")}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── CONSOLIDATED SQLITE & LINE WEBHOOK ENDPOINTS ──────────────────────

# Pydantic schemas for the new booking APIs
class SaveCustomerPayload(BaseModel):
    queueId: str
    name: str
    lineId: str
    astrologer: str
    service: str
    date: str
    slot: str
    questions: str = ""
    birthDate: str = ""
    birthHour: str = ""
    birthMin: str = ""
    birthCountry: str = ""
    birthCity: str = ""
    gcalEventId: str = ""
    lineMessage: str = ""

# LINE Messaging API Helpers
def find_user_id_by_line_id(line_id: str) -> Optional[str]:
    if not line_id:
        return None
    line_id_clean = line_id.strip().lower()
    if line_id_clean.startswith('@'):
        line_id_clean = line_id_clean[1:]
    conn = db.get_db_connection()
    c = conn.cursor()
    c.execute("SELECT user_id FROM customers WHERE LOWER(REPLACE(line_id, '@', '')) = ? ORDER BY seq DESC LIMIT 1", (line_id_clean,))
    row = c.fetchone()
    conn.close()
    if row:
        return row[0]
    return None

def push_line_message(user_id: str, text: str) -> bool:
    if not LINE_CHANNEL_ACCESS_TOKEN or not user_id:
        print(f"[LINE] Skipping push (token or user_id missing)")
        return False
    try:
        url = "https://api.line.me/v2/bot/message/push"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {LINE_CHANNEL_ACCESS_TOKEN}"
        }
        payload = {
            "to": user_id,
            "messages": [{"type": "text", "text": text}]
        }
        res = requests.post(url, headers=headers, json=payload, timeout=5)
        if res.status_code == 200:
            print(f"[LINE] Sent push successfully to {user_id}")
            return True
        else:
            print(f"[LINE] Failed to send push ({res.status_code}): {res.text}")
    except Exception as e:
        print(f"[LINE] Error sending push: {e}")
    return False

PROFILE_CACHE = {}
def get_line_profile(user_id: str) -> dict:
    if user_id in PROFILE_CACHE:
        return PROFILE_CACHE[user_id]
    if not LINE_CHANNEL_ACCESS_TOKEN:
        return {"displayName": "ลูกค้า LINE"}
    try:
        url = f"https://api.line.me/v2/bot/profile/{user_id}"
        headers = {"Authorization": f"Bearer {LINE_CHANNEL_ACCESS_TOKEN}"}
        res = requests.get(url, headers=headers, timeout=5)
        if res.status_code == 200:
            profile = res.json()
            PROFILE_CACHE[user_id] = profile
            return profile
        else:
            print(f"[LINE] Failed to get profile ({res.status_code}): {res.text}")
    except Exception as e:
        print(f"[LINE] Error getting profile: {e}")
    return {"displayName": "ลูกค้า LINE"}

def reply_line_message(reply_token: str, reply_text: str):
    if not LINE_CHANNEL_ACCESS_TOKEN or not reply_token:
        return
    try:
        url = "https://api.line.me/v2/bot/message/reply"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {LINE_CHANNEL_ACCESS_TOKEN}"
        }
        payload = {
            "replyToken": reply_token,
            "messages": [{"type": "text", "text": reply_text}]
        }
        res = requests.post(url, headers=headers, json=payload, timeout=5)
        if res.status_code == 200:
            print(f"[LINE] Sent reply successfully")
        else:
            print(f"[LINE] Failed to send reply ({res.status_code}): {res.text}")
    except Exception as e:
        print(f"[LINE] Error sending reply: {e}")

def verify_line_signature(request_data: bytes, signature: str) -> bool:
    if not LINE_CHANNEL_SECRET:
        return True
    hash = hmac.new(
        LINE_CHANNEL_SECRET.encode('utf-8'),
        request_data,
        hashlib.sha256
    ).digest()
    calculated_signature = base64.b64encode(hash).decode('utf-8')
    return hmac.compare_digest(calculated_signature, signature)

# Calendar checking and booking
def parse_date_slot_to_iso(date_str: str, slot_str: str):
    import re
    date_str = date_str.strip()
    match_yyyy = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$", date_str)
    match_dd = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$", date_str)
    yyyy, mm, dd = None, None, None
    if match_yyyy:
        yyyy, mm, dd = match_yyyy.groups()
    elif match_dd:
        dd, mm, yyyy = match_dd.groups()
    else:
        nums = re.findall(r"\d+", date_str)
        if len(nums) >= 3:
            if len(nums[0]) == 4: yyyy, mm, dd = nums[0], nums[1], nums[2]
            elif len(nums[2]) == 4: dd, mm, yyyy = nums[0], nums[1], nums[2]
    if not yyyy or not mm or not dd:
        raise ValueError("รูปแบบวันที่ไม่ถูกต้อง (ตัวอย่าง: 2026-06-30 หรือ 30/06/2026)")
    year_val = int(yyyy)
    if year_val > 2500: year_val -= 543
    date_normalized = f"{year_val:04d}-{int(mm):02d}-{int(dd):02d}"
    
    slot_str = slot_str.replace(" ", "").replace(".", ":")
    times = re.findall(r"\d{2}:\d{2}", slot_str)
    if len(times) == 2:
        start_time, end_time = times[0], times[1]
    elif len(times) == 1:
        start_time = times[0]
        h, m = map(int, start_time.split(":"))
        end_time = f"{(h+1)%24:02d}:{m:02d}"
    else:
        nums = re.findall(r"\d+", slot_str)
        if len(nums) >= 2:
            start_time = f"{int(nums[0]):02d}:00"
            end_time = f"{int(nums[1]):02d}:00"
        else:
            raise ValueError("รูปแบบเวลานัดไม่ถูกต้อง (ตัวอย่าง: 10:00 - 11:00)")
            
    start_iso = f"{date_normalized}T{start_time}:00+07:00"
    end_iso = f"{date_normalized}T{end_time}:00+07:00"
    return start_iso, end_iso

def format_date_to_dd_mm_yyyy(date_str: str) -> str:
    import re
    date_str = date_str.strip()
    match_yyyy = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$", date_str)
    match_dd = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$", date_str)
    yyyy, mm, dd = None, None, None
    if match_yyyy:
        yyyy, mm, dd = match_yyyy.groups()
    elif match_dd:
        dd, mm, yyyy = match_dd.groups()
    else:
        nums = re.findall(r"\d+", date_str)
        if len(nums) >= 3:
            if len(nums[0]) == 4: yyyy, mm, dd = nums[0], nums[1], nums[2]
            elif len(nums[2]) == 4: dd, mm, yyyy = nums[0], nums[1], nums[2]
    if not yyyy or not mm or not dd:
        return date_str
    year_val = int(yyyy)
    if year_val > 2500: year_val -= 543
    return f"{int(dd):02d}/{int(mm):02d}/{year_val:04d}"

def check_calendar_slot_busy(date_str: str, slot_str: str) -> dict:
    if not GOOGLE_SCRIPT_URL:
        return {"success": True, "busy": False}
    try:
        start_iso, end_iso = parse_date_slot_to_iso(date_str, slot_str)
    except Exception as e:
        return {"success": False, "error": "parse_error", "message": str(e)}
    payload = {
        "action": "checkSlot",
        "startTime": start_iso,
        "endTime": end_iso
    }
    try:
        res = requests.post(GOOGLE_SCRIPT_URL, json=payload, timeout=10)
        if res.status_code == 200:
            return res.json()
        return {"success": False, "error": "http_error", "message": f"HTTP {res.status_code}"}
    except Exception as e:
        return {"success": True, "busy": False, "warning": str(e)}

def check_and_block_calendar_event(date_str: str, slot_str: str, name: str, line_id: str, service: str, astrologer: str, questions: str = "") -> dict:
    if not GOOGLE_SCRIPT_URL:
        return {"success": True, "eventId": f"mock-gcal-{datetime.datetime.now().microsecond}"}
    try:
        start_iso, end_iso = parse_date_slot_to_iso(date_str, slot_str)
    except Exception as e:
        return {"success": False, "error": "parse_error", "message": str(e)}
    summary = f"MuHub: {name} ({astrologer})"
    description = (
        f"คิวจอง: {summary}\n"
        f"แพคเกจ: {service}\n"
        f"ไลน์ลูกค้า: {line_id}\n"
        f"คำถามพิเศษ: {questions}\n"
        f"จองผ่านแชท LINE OA"
    )
    payload = {
        "action": "checkAndCreate",
        "startTime": start_iso,
        "endTime": end_iso,
        "summary": summary,
        "description": description
    }
    try:
        res = requests.post(GOOGLE_SCRIPT_URL, json=payload, timeout=10)
        if res.status_code == 200:
            return res.json()
        return {"success": False, "error": "http_error", "message": f"HTTP {res.status_code}"}
    except Exception as e:
        return {"success": False, "error": "connection_error", "message": str(e)}

def download_line_image(message_id: str) -> Optional[bytes]:
    if not LINE_CHANNEL_ACCESS_TOKEN:
        return None
    url = f"https://api-data.line.me/v2/bot/message/{message_id}/content"
    headers = {"Authorization": f"Bearer {LINE_CHANNEL_ACCESS_TOKEN}"}
    try:
        res = requests.get(url, headers=headers, timeout=15)
        if res.status_code == 200:
            return res.content
    except Exception as e:
        print(f"[LINE Image] Error downloading {message_id}: {e}")
    return None

def verify_slip_via_slipok(image_bytes: bytes, expected_amount: Optional[float] = None) -> dict:
    if not SLIPOK_API_KEY or not SLIPOK_BRANCH_ID:
        return {"success": False, "error": "not_configured", "message": "ไม่ได้ตั้งค่า SlipOK"}
    url = f"https://api.slipok.com/api/line/apikey/{SLIPOK_BRANCH_ID}"
    headers = {"x-authorization": SLIPOK_API_KEY}
    files = {"files": ("slip.png", image_bytes, "image/png")}
    data = {"log": "true"}
    if expected_amount:
        data["amount"] = expected_amount
    try:
        res = requests.post(url, headers=headers, files=files, data=data, timeout=15)
        if res.status_code == 200:
            return res.json()
        return {"success": False, "error": "http_error", "message": f"HTTP {res.status_code}"}
    except Exception as e:
        return {"success": False, "error": "exception", "message": str(e)}

def confirm_and_gcal_booking(queue_id: str) -> dict:
    queue_id_clean = queue_id.strip().upper()
    conn = db.get_db_connection()
    c = conn.cursor()
    
    # 1. Check pending_customers
    c.execute("SELECT * FROM pending_customers WHERE UPPER(queue_id) = ?", (queue_id_clean,))
    p_row = c.fetchone()
    
    if p_row:
        p_cust = dict(p_row)
        cal_result = check_and_block_calendar_event(
            date_str=p_cust["booking_date_appointment"],
            slot_str=p_cust["booking_slot"],
            name=p_cust["customer_name"],
            line_id=p_cust["line_id"],
            service=p_cust["package"],
            astrologer=p_cust["astrologer"],
            questions=p_cust["questions"]
        )
        if not cal_result.get("success"):
            conn.close()
            return {"success": False, "message": f"❌ ตารางเวลาชนกัน หรือล็อกปฏิทินไม่สำเร็จ: {cal_result.get('message')}"}
            
        gcal_event_id = cal_result.get("eventId")
        
        # Determine client booking count and type
        c.execute("SELECT booking_count FROM customers WHERE user_id = ?", (p_cust["user_id"],))
        old_count = c.fetchone()
        new_count = (old_count[0] + 1) if old_count else 1
        
        c.execute("DELETE FROM customers WHERE user_id = ?", (p_cust["user_id"],))
        c.execute("""
        INSERT INTO customers (
            timestamp, display_name, user_id, status, booking_date, customer_name, line_id,
            birth_date, birth_time, birth_country, birth_city, package, astrologer,
            booking_date_appointment, booking_slot, questions, queue_id, payment_status,
            customer_type, booking_count, notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            p_cust["timestamp"], p_cust["display_name"], p_cust["user_id"], "ยืนยัน",
            p_cust["booking_date"], p_cust["customer_name"], p_cust["line_id"],
            p_cust["birth_date"], p_cust["birth_time"], p_cust["birth_country"], p_cust["birth_city"],
            p_cust["package"], p_cust["astrologer"], p_cust["booking_date_appointment"],
            p_cust["booking_slot"], p_cust["questions"], p_cust["queue_id"], "ชำระแล้ว",
            "ลูกค้าเก่า" if new_count > 1 else "ลูกค้าใหม่", new_count,
            f"ยืนยันคิวและล็อก GCal สำเร็จ (Event ID: {gcal_event_id})"
        ))
        
        c.execute("DELETE FROM pending_customers WHERE UPPER(queue_id) = ?", (queue_id_clean,))
        conn.commit()
        conn.close()
        
        # Send confirmation
        user_id = p_cust["user_id"]
        if user_id:
            booking_date_formatted = format_date_to_dd_mm_yyyy(p_cust["booking_date_appointment"])
            reply_text = (
                f"✨ ยืนยันการจองคิวสำเร็จแล้วค่ะ คุณ {p_cust['customer_name']}! 🙏✨\n\n"
                f"ทางทีมงานได้ตรวจสอบสลิปการโอนเงินเรียบร้อยแล้วค่ะ และได้ลงคิวนัดหมายของคุณเรียบร้อยแล้วนะคะ 📝\n\n"
                f"📋 รายละเอียดคิวของท่าน:\n"
                f"• เลขอ้างอิง: {queue_id_clean}\n"
                f"• แพคเกจ: {p_cust['package']}\n"
                f"• หมอดู: {p_cust['astrologer']}\n"
                f"• วันนัดหมาย: {booking_date_formatted}\n"
                f"• ช่วงเวลานัด: {p_cust['booking_slot']}\n\n"
                f"📞 เมื่อถึงเวลานัดหมาย จะมีการติดต่อกลับโดยการโทรทาง LINE OA นะคะ\n\n"
                f"ขอให้มีช่วงเวลาทำนายดวงที่ดีนะคะ! ขอบคุณค่ะ 💖"
            )
            push_line_message(user_id, reply_text)
        return {"success": True, "message": f"✅ ยืนยันคิว {queue_id_clean} สำเร็จ"}

    # 2. Check bookings table
    c.execute("SELECT * FROM bookings WHERE UPPER(queue_id) = ?", (queue_id_clean,))
    b_row = c.fetchone()
    
    if b_row:
        booking = dict(b_row)
        if booking["status"] == "ชำระแล้ว" and booking["gcal_event_id"]:
            conn.close()
            return {"success": False, "message": f"คิว {queue_id_clean} ได้ยืนยันเรียบร้อยแล้วก่อนหน้านี้ค่ะ"}
            
        cal_result = check_and_block_calendar_event(
            date_str=booking["date"],
            slot_str=booking["slot"],
            name=booking["name"],
            line_id=booking["line_id"],
            service=booking["service"],
            astrologer=booking["astrologer"],
            questions=booking["questions"]
        )
        if not cal_result.get("success"):
            conn.close()
            return {"success": False, "message": f"❌ ตารางเวลาชนกัน หรือล็อกปฏิทินไม่สำเร็จ: {cal_result.get('message')}"}
            
        gcal_event_id = cal_result.get("eventId")
        c.execute("UPDATE bookings SET status = 'ชำระแล้ว', gcal_event_id = ? WHERE UPPER(queue_id) = ?", (gcal_event_id, queue_id_clean))
        conn.commit()
        conn.close()
        
        user_id = find_user_id_by_line_id(booking["line_id"])
        if user_id:
            booking_date_formatted = format_date_to_dd_mm_yyyy(booking["date"])
            reply_text = (
                f"✨ ยืนยันการจองคิวสำเร็จแล้วค่ะ คุณ {booking['name']}! 🙏✨\n\n"
                f"ทางทีมงานได้ตรวจสอบสลิปการโอนเงินเรียบร้อยแล้วค่ะ และได้ลงคิวนัดหมายของคุณเรียบร้อยแล้วนะคะ 📝\n\n"
                f"📋 รายละเอียดคิวของท่าน:\n"
                f"• เลขอ้างอิง: {queue_id_clean}\n"
                f"• แพคเกจ: {booking['service']}\n"
                f"• หมอดู: {booking['astrologer']}\n"
                f"• วันนัดหมาย: {booking_date_formatted}\n"
                f"• ช่วงเวลานัด: {booking['slot']}\n\n"
                f"📞 เมื่อถึงเวลานัดหมาย จะมีการติดต่อกลับโดยการโทรทาง LINE OA นะคะ\n\n"
                f"ขอให้มีช่วงเวลาทำนายดวงที่ดีนะคะ! ขอบคุณค่ะ 💖"
            )
            push_line_message(user_id, reply_text)
        return {"success": True, "message": f"✅ ยืนยันคิว {queue_id_clean} สำเร็จ"}

    conn.close()
    return {"success": False, "message": f"ไม่พบรหัสคิว {queue_id_clean} ในตารางฐานข้อมูลที่รอยืนยันค่ะ"}

def parse_booking_message(text: str) -> dict:
    fields = {
        "name": "", "line_id": "", "birth_date": "", "birth_time": "",
        "country": "", "city": "", "questions": "", "queue_id": "",
        "package": "", "astrologer": "", "booking_date": "", "booking_slot": "",
        "payment_status": ""
    }
    lines = text.split("\n")
    for line in lines:
        line = line.strip()
        if not line.startswith("•"):
            continue
        parts = line.split(":", 1)
        if len(parts) < 2:
            continue
        key = parts[0].replace("•", "").strip()
        val = parts[1].strip()
        
        if "ชื่อ" in key: fields["name"] = val
        elif "Line ID" in key or "ไลน์" in key: fields["line_id"] = val
        elif "วันเกิด" in key: fields["birth_date"] = val
        elif "เวลาเกิด" in key: fields["birth_time"] = val
        elif "ประเทศ" in key: fields["country"] = val
        elif "จังหวัด" in key: fields["city"] = val
        elif "คำถามพิเศษ" in key: fields["questions"] = val
        elif "เลขอ้างอิง" in key: fields["queue_id"] = val
        elif "แพคเกจ" in key: fields["package"] = val
        elif "หมอดู" in key: fields["astrologer"] = val
        elif "วันนัด" in key: fields["booking_date"] = val
        elif "เวลานัด" in key: fields["booking_slot"] = val
        elif "สถานะการชำระเงิน" in key: fields["payment_status"] = val
    return fields

# FastAPI Endpoint definitions
@app.post("/api/save-customer")
async def save_customer(payload: SaveCustomerPayload, request: Request):
    verify_app_token_and_rate_limit(request, "save_customer")
    
    print(f"[Direct Web API] Saving booking: {payload.name} / {payload.queueId}")
    try:
        conn = db.get_db_connection()
        c = conn.cursor()
        c.execute("""
        INSERT OR REPLACE INTO bookings (
            timestamp, queue_id, name, line_id, astrologer, service, date, slot, questions,
            birth_date, birth_hour, birth_min, birth_country, birth_city, gcal_event_id, status
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            payload.queueId, payload.name, payload.lineId, payload.astrologer, payload.service,
            payload.date, payload.slot, payload.questions, payload.birthDate, payload.birthHour,
            payload.birthMin, payload.birthCountry, payload.birthCity, payload.gcalEventId, "pending"
        ))
        conn.commit()
        conn.close()
        
        # LINE Notifications
        user_id = find_user_id_by_line_id(payload.lineId)
        line_sent = False
        if user_id and payload.lineMessage:
            line_sent = push_line_message(user_id, payload.lineMessage)
            
        if LINE_ADMIN_USER_ID and payload.lineMessage:
            admin_text = f"🔔 [ระบบจองคิว] มีคิวใหม่จองเข้ามาค่ะ!\n\n{payload.lineMessage}"
            push_line_message(LINE_ADMIN_USER_ID, admin_text)
            
        # Export DB to local Excel sheet automatically to sync G Drive (if running locally)
        if os.path.exists(os.path.dirname(EXCEL_PATH)):
            try:
                db.export_sqlite_to_excel(EXCEL_PATH)
            except Exception as e:
                print(f"[Excel Sync] Auto sync Excel failed: {e}")
                
        return {"success": True, "queueId": payload.queueId, "lineSent": line_sent}
    except Exception as e:
        print(f"[Direct Web API] Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/webhook")
@app.post("/api/line-webhook")
async def line_webhook(request: Request):
    signature = request.headers.get("X-Line-Signature", "")
    body_bytes = await request.body()
    
    if not verify_line_signature(body_bytes, signature):
        print("[Webhook] Warning: Invalid signature, request rejected.")
        raise HTTPException(status_code=400, detail="Invalid signature")
        
    try:
        payload = json.loads(body_bytes.decode("utf-8"))
    except Exception as e:
        raise HTTPException(status_code=400, detail="Bad JSON format")
        
    events = payload.get("events", [])
    print(f"[Webhook] Received LINE events: {len(events)}")
    
    for event in events:
        event_type = event.get("type")
        source = event.get("source", {})
        source_type = source.get("type")
        
        if event_type != "message" or source_type != "user":
            continue
            
        message = event.get("message", {})
        message_type = message.get("type")
        
        if message_type not in ["text", "image"]:
            continue
            
        user_id = source.get("userId")
        reply_token = event.get("replyToken")
        profile = get_line_profile(user_id)
        display_name = profile.get("displayName", "ลูกค้า LINE")
        
        conn = db.get_db_connection()
        c = conn.cursor()
        
        if message_type == "image":
            message_id = message.get("id")
            print(f"[Webhook] Image received (Message ID: {message_id})")
            
            # Find in pending_customers
            c.execute("SELECT * FROM pending_customers WHERE user_id = ? ORDER BY seq DESC LIMIT 1", (user_id,))
            p_row = c.fetchone()
            
            if not p_row:
                print(f"[Webhook] No pending booking for user_id: {user_id}")
                conn.close()
                continue
                
            p_cust = dict(p_row)
            queue_id = p_cust["queue_id"]
            customer_name = p_cust["customer_name"]
            conn.close()
            
            if not SLIPOK_API_KEY or not SLIPOK_BRANCH_ID:
                print(f"[Webhook] SlipOK not set. Admin alert.")
                if LINE_ADMIN_USER_ID:
                    admin_notify = (
                        f"🔔 [แจ้งโอนเงิน] คุณ {customer_name} ส่งรูปสลิปเข้ามาแล้วสำหรับคิว {queue_id} ค่ะ\n"
                        f"กรุณาตรวจสอบสลิปแล้วพิมพ์:\n"
                        f"👉 ยืนยัน {queue_id}"
                    )
                    push_line_message(LINE_ADMIN_USER_ID, admin_notify)
                reply_line_message(
                    reply_token,
                    f"🙏 ได้รับหลักฐานการโอนเงินเรียบร้อยแล้วค่ะ คุณ {customer_name}!\n\n"
                    f"แอดมินจะรีบทำการตรวจสอบสลิปและแจ้งเตือนยืนยันคิวให้ท่านโดยเร็วที่สุดนะคะ 📝"
                )
                continue
                
            # SlipOK checking
            image_bytes = download_line_image(message_id)
            if not image_bytes:
                reply_line_message(
                    reply_token,
                    "❌ ไม่สามารถดาวน์โหลดรูปภาพสลิปได้ชั่วคราว รบกวนลองส่งสลิปใหม่อีกครั้งค่ะ"
                )
                continue
                
            res_slip = verify_slip_via_slipok(image_bytes)
            if res_slip.get("success") and res_slip.get("data", {}).get("success"):
                slip_data = res_slip["data"]
                amount = slip_data.get("amount", 0)
                trans_ref = slip_data.get("transRef", "")
                
                confirm_res = confirm_and_gcal_booking(queue_id)
                if confirm_res.get("success"):
                    if LINE_ADMIN_USER_ID:
                        admin_notify = (
                            f"✅ [ตรวจสอบสลิปอัตโนมัติ] คุณ {customer_name} (คิว: {queue_id}) ผ่าน SlipOK แล้วค่ะ\n"
                            f"• ยอดโอน: {amount} บาท\n"
                            f"• รหัสอ้างอิง: {trans_ref}\n"
                            f"• ลงปฏิทินและบันทึกฐานข้อมูลหลักสำเร็จแล้วค่ะ"
                        )
                        push_line_message(LINE_ADMIN_USER_ID, admin_notify)
                else:
                    err_msg = confirm_res.get("message", "Error")
                    reply_line_message(
                        reply_token,
                        f"⚠️ ตรวจสอบสลิปผ่านแล้ว แต่ระบบพบล็อกเวลาไม่ว่างหรือขัดข้อง: {err_msg}\n"
                        f"ไม่ต้องเป็นกังวลนะคะ แอดมินกำลังเร่งจองคิวให้ท่านแมนนวลค่ะ 🙏"
                    )
            else:
                err_msg = res_slip.get("message") or "ไม่พบ QR Code หรือเป็นสลิปซ้ำ"
                reply_line_message(
                    reply_token,
                    f"❌ ระบบไม่สามารถตรวจสอบข้อมูลสลิปอัตโนมัติ ({err_msg}) 😢\n"
                    f"แอดมินจะทำการตรวจสอบสลิปโอนเงินของคุณด้วยตนเองเร็วๆ นี้ค่ะ 🙏"
                )
            continue
            
        # Text Message
        message_text = message.get("text", "")
        
        # Check Admin commands
        is_admin_cmd = (user_id == LINE_ADMIN_USER_ID) and (
            message_text.strip().startswith("ยืนยัน ") or 
            message_text.strip().lower().startswith("/confirm ")
        )
        if is_admin_cmd:
            parts = message_text.strip().split(" ", 1)
            queue_id_cmd = parts[1].strip() if len(parts) > 1 else ""
            result = confirm_and_gcal_booking(queue_id_cmd)
            reply_line_message(reply_token, result["message"])
            conn.close()
            return "OK"
            
        # Check if booking template
        is_booking = "📋 ข้อมูลวันเกิดสำหรับทำนายดวง" in message_text
        if is_booking:
            booking_data = parse_booking_message(message_text)
            cal_check = check_calendar_slot_busy(
                date_str=booking_data.get("booking_date", ""),
                slot_str=booking_data.get("booking_slot", "")
            )
            if not cal_check.get("success") and cal_check.get("error") == "parse_error":
                reply_line_message(
                    reply_token,
                    f"ขออภัยค่ะ รูปแบบวันที่หรือเวลานัดที่ระบุไม่ถูกต้อง: {cal_check.get('message')} 😢\n"
                    f"รบกวนส่งแบบฟอร์มอีกครั้งนะคะ"
                )
                conn.close()
                return "OK"
                
            if cal_check.get("busy"):
                reply_line_message(
                    reply_token,
                    f"ขออภัยค่ะ คุณ {booking_data.get('name')}! 🙏\n"
                    f"วันและเวลานัดหมายที่คุณเลือกมีผู้อื่นจองไว้เรียบร้อยแล้วค่ะ 😢"
                )
                conn.close()
                return "OK"
                
            # Save to pending_customers
            c.execute("DELETE FROM pending_customers WHERE user_id = ?", (user_id,))
            c.execute("""
            INSERT INTO pending_customers (
                timestamp, display_name, user_id, status, booking_date, customer_name, line_id,
                birth_date, birth_time, birth_country, birth_city, package, astrologer,
                booking_date_appointment, booking_slot, questions, queue_id, payment_status,
                customer_type, booking_count, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
                display_name, user_id, "ใหม่", datetime.datetime.now().strftime("%d/%m/%Y"),
                booking_data.get("name", ""), booking_data.get("line_id", ""),
                booking_data.get("birth_date", ""), booking_data.get("birth_time", ""),
                booking_data.get("country", ""), booking_data.get("city", ""),
                booking_data.get("package", ""), booking_data.get("astrologer", ""),
                booking_data.get("booking_date", ""), booking_data.get("booking_slot", ""),
                booking_data.get("questions", ""), booking_data.get("queue_id", ""),
                "รอตรวจสอบ", "ลูกค้าใหม่", "จองผ่านแชท LINE (รอสลิป)"
            ))
            conn.commit()
            
            booking_date_formatted = format_date_to_dd_mm_yyyy(booking_data.get('booking_date', ''))
            reply_text = (
                f"ได้รับข้อมูลการจองคิวเรียบร้อยแล้วค่ะ คุณ {booking_data.get('name')}! 🙏✨\n\n"
                f"📋 รายละเอียดการจองคิวของท่าน:\n"
                f"• เลขอ้างอิง: {booking_data.get('queue_id')}\n"
                f"• แพคเกจ: {booking_data.get('package')}\n"
                f"• หมอดู: {booking_data.get('astrologer')}\n"
                f"• วันนัดหมาย: {booking_date_formatted}\n"
                f"• ช่วงเวลานัด: {booking_data.get('booking_slot')}\n\n"
                f"💵 ช่วงเวลานี้ว่างและจองได้ค่ะ รบกวนส่งสลิปโอนเงินเข้ามาเพื่อยืนยันสลอตเวลานะคะ 💖"
            )
            reply_line_message(reply_token, reply_text)
            conn.close()
            return "OK"
            
        # Chat responses and logs
        booking_keywords = ["จอง", "ดูดวง", "คิว", "package", "แพคเกจ", "หมอดู", "ผูกดวง", "ทำนาย", "จองคิว"]
        is_asking_booking = any(kw in message_text.lower() for kw in booking_keywords)
        
        c.execute("SELECT seq, notes FROM customers WHERE user_id = ?", (user_id,))
        cust_row = c.fetchone()
        
        if cust_row and not is_asking_booking:
            # Append notes for existing user
            old_notes = cust_row[1] or ""
            timestamp = datetime.datetime.now().strftime("%d/%m %H:%M")
            new_notes = f"{old_notes}\n[{timestamp}] ลูกค้า: {message_text}".strip()
            c.execute("UPDATE customers SET notes = ? WHERE user_id = ?", (new_notes, user_id))
            conn.commit()
        else:
            if not cust_row:
                # Insert customer as a lead
                c.execute("""
                INSERT INTO customers (
                    timestamp, display_name, user_id, status, notes
                ) VALUES (?, ?, ?, ?, ?)
                """, (
                    datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
                    display_name, user_id, "ผู้สนใจใหม่", f"ทักแชทครั้งแรก: {message_text}"
                ))
                conn.commit()
            
            # Send intro/booking guide
            reply_text = (
                f"สวัสดีค่ะ คุณ {display_name} ยินดีต้อนรับสู่ MuHub ค่ะ! 🙏✨\n\n"
                f"หากคุณต้องการจองคิวดูดวง/ผูกดวงชะตา สามารถทำรายการได้ 2 ช่องทางค่ะ:\n\n"
                f"🌟 ช่องทางที่ 1: จองออนไลน์ด้วยตนเอง (รวดเร็ว ทราบคิวว่างทันที)\n"
                f"👉 ลิงก์: {WEB_APP_URL}\n\n"
                f"📝 ช่องทางที่ 2: จองผ่านแชทนี้\n"
                f"กรุณากรอกแบบฟอร์มด้านล่าง ส่งเข้ามาพร้อมสลิปโอนเงินนะคะ:\n\n"
                f"📋 ข้อมูลวันเกิดสำหรับทำนายดวง\n"
                f"• ชื่อ:\n"
                f"• Line ID:\n"
                f"• วันเกิด:\n"
                f"• เวลาเกิด:\n"
                f"• ประเทศ: ประเทศไทย\n"
                f"• จังหวัด:\n"
                f"• คำถามพิเศษ:\n\n"
                f"• เลขอ้างอิง: (ระบุ 'จองผ่านแชท')\n"
                f"• แพคเกจ:\n"
                f"• หมอดู:\n"
                f"• วันนัด:\n"
                f"• เวลานัด:"
            )
            reply_line_message(reply_token, reply_text)
            
        conn.close()
        
    # Auto-export local excel sync
    if os.path.exists(os.path.dirname(EXCEL_PATH)):
        try:
            db.export_sqlite_to_excel(EXCEL_PATH)
        except Exception as e:
            print(f"[Excel Sync] Auto sync Excel failed: {e}")
            
    return "OK"

@app.get("/api/check-promo")
def check_promo(code: str, lineId: str):
    code_clean = code.strip().upper()
    line_clean = lineId.strip().lower()
    if not code_clean or not line_clean:
        raise HTTPException(status_code=400, detail="Missing parameter code or lineId")
        
    conn = db.get_db_connection()
    c = conn.cursor()
    c.execute("SELECT 1 FROM promo_used WHERE code = ? AND line_id = ?", (code_clean, line_clean))
    used = c.fetchone()
    conn.close()
    
    return {"eligible": not bool(used), "code": code_clean}

@app.post("/api/redeem-promo")
async def redeem_promo(payload: Dict[str, Any], request: Request):
    verify_app_token_and_rate_limit(request, "redeem_promo")
    code = str(payload.get("code", "")).strip().upper()
    line_id = str(payload.get("lineId", "")).strip().lower()
    if not code or not line_id:
        raise HTTPException(status_code=400, detail="Missing code or lineId")
        
    conn = db.get_db_connection()
    c = conn.cursor()
    try:
        c.execute("INSERT INTO promo_used (code, line_id) VALUES (?, ?)", (code, line_id))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return {"success": False, "reason": "already_used"}
        
    conn.close()
    return {"success": True}

@app.get("/api/booking-status")
def booking_status(queueId: str):
    queue_id_clean = queueId.strip()
    if not queue_id_clean:
        raise HTTPException(status_code=400, detail="Missing queueId")
        
    conn = db.get_db_connection()
    c = conn.cursor()
    
    # 1. Check bookings table
    c.execute("SELECT status, gcal_event_id FROM bookings WHERE queue_id = ?", (queue_id_clean,))
    b_row = c.fetchone()
    if b_row:
        is_confirmed = b_row["status"] == "ชำระแล้ว" or bool(b_row["gcal_event_id"])
        conn.close()
        return {
            "found": True,
            "sheet": "Bookings",
            "status": "confirmed" if is_confirmed else "pending",
            "raw_status": b_row["status"]
        }
        
    # 2. Check customers (confirmed LINE customers)
    c.execute("SELECT status, payment_status FROM customers WHERE queue_id = ?", (queue_id_clean,))
    c_row = c.fetchone()
    if c_row:
        is_confirmed = c_row["payment_status"] == "ชำระแล้ว" or c_row["status"] == "ยืนยัน"
        conn.close()
        return {
            "found": True,
            "sheet": "Cust",
            "status": "confirmed" if is_confirmed else "pending",
            "raw_status": c_row["payment_status"] or c_row["status"]
        }
        
    # 3. Check pending_customers
    c.execute("SELECT status, payment_status FROM pending_customers WHERE queue_id = ?", (queue_id_clean,))
    p_row = c.fetchone()
    if p_row:
        is_confirmed = p_row["payment_status"] == "ชำระแล้ว" or p_row["status"] == "ยืนยัน"
        conn.close()
        return {
            "found": True,
            "sheet": "PendingCust",
            "status": "confirmed" if is_confirmed else "pending",
            "raw_status": p_row["payment_status"] or p_row["status"]
        }
        
    conn.close()
    return {"found": False}

@app.post("/api/track-event")
async def track_event(payload: Dict[str, Any], request: Request):
    verify_app_token_and_rate_limit(request, "track_event")
    event_name = str(payload.get("event", ""))[:100].strip()
    session_id = str(payload.get("sessionId", ""))[:100].strip()
    detail = json.dumps(payload.get("detail", {}), ensure_ascii=False)
    
    if not event_name:
        raise HTTPException(status_code=400, detail="Missing event name")
        
    try:
        conn = db.get_db_connection()
        c = conn.cursor()
        c.execute("INSERT INTO events_log (timestamp, event_name, session_id, detail) VALUES (?, ?, ?, ?)",
                  (datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), event_name, session_id, detail))
        conn.commit()
        conn.close()
        return {"success": True}
    except Exception as e:
        print(f"[TrackEvent] Error: {e}")
        return {"success": False, "error": str(e)}

@app.get("/api/health")
def health():
    excel_exists = os.path.exists(EXCEL_PATH)
    return {
        "status": "ok",
        "sqlite_db_exists": os.path.exists(db.DB_PATH),
        "excel_exists": excel_exists,
        "excel_path": EXCEL_PATH,
        "line_token_set": bool(LINE_CHANNEL_ACCESS_TOKEN),
        "line_secret_set": bool(LINE_CHANNEL_SECRET)
    }

# ── EXPORT / IMPORT EXCEL API (For Admin / Cloud Use) ─────────────────
@app.get("/api/admin/export")
def admin_export_excel(token: str = Query(None)):
    if APP_API_TOKEN and token != APP_API_TOKEN:
        raise HTTPException(status_code=401, detail="Unauthorized")
        
    temp_excel = os.path.join(BASE_DIR, "MuHub_Customer_DB_export.xlsx")
    try:
        db.export_sqlite_to_excel(temp_excel)
        return FileResponse(temp_excel, filename="MuHub_Customer_DB.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

@app.post("/api/admin/import")
async def admin_import_excel(file: UploadFile = File(...), token: str = Query(None)):
    if APP_API_TOKEN and token != APP_API_TOKEN:
        raise HTTPException(status_code=401, detail="Unauthorized")
        
    temp_path = os.path.join(BASE_DIR, "MuHub_Customer_DB_uploaded.xlsx")
    try:
        with open(temp_path, "wb") as buffer:
            buffer.write(await file.read())
            
        # Parse Excel file and verify if any row transitioned to approved
        # This mirrors the background monitoring thread auto-confirmations.
        wb = openpyxl.load_workbook(temp_path, data_only=True)
        
        approved_count = 0
        
        # 1. Verify Bookings auto-confirms in uploaded sheet
        if "Bookings" in wb.sheetnames:
            ws = wb["Bookings"]
            # Find status column from header
            col_gcal = 15
            col_status = 16
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col_idx).value
                if val == "GCal Event ID": col_gcal = col_idx
                elif val == "Status": col_status = col_idx
                
            for r in range(2, ws.max_row + 1):
                q_id = ws.cell(row=r, column=2).value
                status_val = ws.cell(row=r, column=col_status).value
                gcal_id = ws.cell(row=r, column=col_gcal).value
                
                if q_id and status_val == "ชำระแล้ว" and not gcal_id:
                    # Trigger confirmation logic
                    confirm_res = confirm_and_gcal_booking(str(q_id))
                    if confirm_res.get("success"):
                        approved_count += 1
                        
        # 2. Verify PendingCust auto-confirms in uploaded sheet
        if "PendingCust" in wb.sheetnames:
            ws = wb["PendingCust"]
            for r in range(3, ws.max_row + 1):
                q_id = ws.cell(row=r, column=18).value
                status_val = ws.cell(row=r, column=5).value
                pay_status_val = ws.cell(row=r, column=19).value
                
                is_confirmed = (status_val == "ยืนยัน") or (pay_status_val == "ชำระแล้ว")
                if q_id and is_confirmed:
                    confirm_res = confirm_and_gcal_booking(str(q_id))
                    if confirm_res.get("success"):
                        approved_count += 1
                        
        wb.close()
        
        # Re-run full database synchronization from this newly uploaded sheet
        db.import_excel_to_sqlite(temp_path)
        
        # Overwrite local EXCEL_PATH if we are locally running and it exists
        if os.path.exists(os.path.dirname(EXCEL_PATH)):
            try:
                if os.path.exists(EXCEL_PATH):
                    os.remove(EXCEL_PATH)
                os.rename(temp_path, EXCEL_PATH)
            except Exception as e:
                print(f"[Admin Import] Failed to update local Excel file: {e}")
        else:
            if os.path.exists(temp_path):
                os.remove(temp_path)
                
        return {"success": True, "message": "Import completed successfully", "auto_confirmed_slots": approved_count}
    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

# ── LOCAL EXCEL SYNC & MONITOR THREAD ───────────────────────────────
# Scans Excel files locally for updates, mirrors changes to SQLite, and triggers auto-confirms.
def check_excel_modifications_and_sync():
    if not os.path.exists(EXCEL_PATH):
        return
        
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        changes_detected = False
        
        # 1. Scan PendingCust for manual updates by admin
        if "PendingCust" in wb.sheetnames:
            ws = wb["PendingCust"]
            for r in range(3, ws.max_row + 1):
                status_val = ws.cell(row=r, column=5).value
                pay_status_val = ws.cell(row=r, column=19).value
                q_id = ws.cell(row=r, column=18).value
                
                is_confirmed = (status_val == "ยืนยัน") or (pay_status_val == "ชำระแล้ว")
                if q_id and is_confirmed:
                    confirm_res = confirm_and_gcal_booking(str(q_id))
                    if confirm_res.get("success"):
                        changes_detected = True
                        
        # 2. Scan Bookings for manual updates
        if "Bookings" in wb.sheetnames:
            ws = wb["Bookings"]
            # Find status/gcal columns
            col_gcal = 15
            col_status = 16
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col_idx).value
                if val == "GCal Event ID": col_gcal = col_idx
                elif val == "Status": col_status = col_idx
                
            for r in range(2, ws.max_row + 1):
                q_id = ws.cell(row=r, column=2).value
                status_val = ws.cell(row=r, column=col_status).value
                gcal_id = ws.cell(row=r, column=col_gcal).value
                
                if q_id and status_val == "ชำระแล้ว" and not gcal_id:
                    confirm_res = confirm_and_gcal_booking(str(q_id))
                    if confirm_res.get("success"):
                        changes_detected = True
                        
        wb.close()
        
        # Re-import modified Excel file into SQLite to ensure full parity
        db.import_excel_to_sqlite(EXCEL_PATH)
        
        # Rewrite the Excel sheet to ensure any updated SQLite state (like moved lines) gets exported
        if changes_detected:
            db.export_sqlite_to_excel(EXCEL_PATH)
            print("[Monitor] Excel sync and auto confirmations successfully applied.")
            
    except Exception as e:
        print(f"[Monitor] Error scanning Excel updates: {e}")

def local_excel_monitor_thread():
    print("[Monitor] Local Excel Sync Thread started.")
    last_mtime = 0
    if os.path.exists(EXCEL_PATH):
        try:
            last_mtime = os.path.getmtime(EXCEL_PATH)
        except Exception:
            last_mtime = 0
            
    while True:
        try:
            time.sleep(10)
            if not os.path.exists(EXCEL_PATH):
                continue
                
            try:
                current_mtime = os.path.getmtime(EXCEL_PATH)
            except Exception:
                continue
                
            if current_mtime != last_mtime:
                print(f"[Monitor] Local Excel file update detected on G Drive.")
                time.sleep(3) # Wait for file lock release
                check_excel_modifications_and_sync()
                try:
                    last_mtime = os.path.getmtime(EXCEL_PATH)
                except Exception:
                    pass
        except Exception as e:
            print(f"[Monitor] Error in Excel monitoring loop: {e}")

# Start local monitor thread on startup
if os.path.exists(os.path.dirname(EXCEL_PATH)):
    monitor_thread = threading.Thread(target=local_excel_monitor_thread, daemon=True)
    monitor_thread.start()

# ── SERVE STATIC FILES ─────────────────────────────────────────────────
@app.get("/")
def get_index():
    index_path = os.path.join(BASE_DIR, "index.html")
    if os.path.exists(index_path):
        return FileResponse(index_path)
    return FileResponse(os.path.join(BASE_DIR, "index.html"))

app.mount("/", StaticFiles(directory=BASE_DIR), name="static")

if __name__ == "__main__":
    import uvicorn
    # Start unified server on unified port
    uvicorn.run(app, host="127.0.0.1", port=PORT)
