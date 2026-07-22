from flask import Flask, request, jsonify
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

app = Flask(__name__)

# ============================================================
# ตั้งค่า
# ============================================================
CHANNEL_ACCESS_TOKEN = "JizxVq097GEGXOe8M6uE3Ixu547xs/v4agkUx3ldk5ffEpsdiug58WmGjjMGestVC6aKOpkdN6pympn8AuOeypNHyVAZq2jjt//US9Vgmm9LwAq5q9kstNXDcAucfo4JZYUVYRnMYWXJ9Ys1VMBvjAdB04t89/1O/w1cDnyilFU="
EXCEL_PATH = r"G:\My Drive\Customer\MuHub_Customer_DB.xlsx"
SHEET_NAME = "Cust"
# ============================================================

HEADERS = {
    "Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}",
    "Content-Type": "application/json"
}

COLOR_HEADER_BG   = "3E2723"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_ROW_ODD     = "FFF8F5"
COLOR_ROW_EVEN    = "FFFFFF"
COLOR_BORDER      = "D7B899"
COLOR_TITLE_BG    = "6D4C41"

COLUMNS = [
    ("ลำดับ",        8),
    ("วันที่บันทึก", 18),
    ("Display Name", 22),
    ("LINE User ID", 28),
    ("สถานะ",        14),
    ("หมายเหตุ",     30),
]

def make_border(color=COLOR_BORDER):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)

def style_data_cell(cell, row_parity):
    bg = COLOR_ROW_ODD if row_parity % 2 == 0 else COLOR_ROW_EVEN
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Angsana New", size=13)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = make_border()

def get_or_create_sheet(wb):
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        ws.row_dimensions[1].height = 30
        ws.merge_cells("A1:F1")
        tc = ws["A1"]
        tc.value = "LINE User IDs — MuHub Customer"
        tc.fill = PatternFill("solid", fgColor=COLOR_TITLE_BG)
        tc.font = Font(name="Angsana New", bold=True, color="FFFFFF", size=16)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 28
        for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = col_name
            cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
            cell.font = Font(name="Angsana New", bold=True, color=COLOR_HEADER_FONT, size=14)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = make_border("5D4037")
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width
        ws.freeze_panes = "A3"
    else:
        ws = wb[SHEET_NAME]
    return ws

def save_user_to_excel(user_id, display_name):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = get_or_create_sheet(wb)

    existing_ids = [row[3] for row in ws.iter_rows(min_row=3, values_only=True) if row[3]]
    if user_id in existing_ids:
        print(f"ℹ️  มีอยู่แล้ว: {display_name} → {user_id}")
        wb.close()
        return

    next_row = max(ws.max_row + 1, 3)
    seq = next_row - 2
    values = [seq, datetime.now().strftime("%d/%m/%Y %H:%M"), display_name, user_id, "ใหม่", ""]
    ws.row_dimensions[next_row].height = 22
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=val)
        style_data_cell(cell, seq)

    wb.save(EXCEL_PATH)
    print(f"✅ บันทึกแล้ว: {display_name} → {user_id}")

def get_display_name(user_id):
    res = requests.get(f"https://api.line.me/v2/bot/profile/{user_id}", headers=HEADERS)
    if res.status_code == 200:
        return res.json().get("displayName", "Unknown")
    return "Unknown"

def send_reply(reply_token, message):
    requests.post("https://api.line.me/v2/bot/message/reply", headers=HEADERS,
                  json={"replyToken": reply_token, "messages": [{"type": "text", "text": message}]})

@app.route("/webhook", methods=["POST"])
def webhook():
    body = request.get_json()
    if not body or "events" not in body:
        return jsonify({"status": "ok"})
    for event in body["events"]:
        event_type = event.get("type")
        user_id = event.get("source", {}).get("userId")
        if not user_id:
            continue
        display_name = get_display_name(user_id)
        save_user_to_excel(user_id, display_name)
        if event_type == "message":
            reply_token = event.get("replyToken")
            if reply_token:
                send_reply(reply_token,
                    f"สวัสดีค่ะ คุณ{display_name} 🌙\nขอบคุณที่ติดต่อ MuHub นะคะ\nทีมงานจะติดต่อกลับเร็วๆ นี้ค่ะ 🙏")
        print(f"📩 Event: {event_type} | User: {display_name} | ID: {user_id}")
    return jsonify({"status": "ok"})

@app.route("/", methods=["GET"])
def home():
    return "MuHub Webhook is running! ✅"

if __name__ == "__main__":
    print("🌙 MuHub Webhook Server เริ่มทำงานแล้วค่ะ...")
    print(f"📁 บันทึก User IDs ที่: {EXCEL_PATH} → sheet: {SHEET_NAME}")
    app.run(port=5000, debug=True)
